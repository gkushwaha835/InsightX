from __future__ import annotations

import csv
import hashlib
import json
import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _normalize(text: str) -> str:
    return ''.join(ch.lower() for ch in (text or '') if ch.isalnum())


def _to_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return 0.0
        return float(value)
    text = str(value).strip().replace(',', '')
    if text == '':
        return 0.0
    if text.endswith('%'):
        text = text[:-1].strip()
    try:
        return float(text)
    except ValueError:
        return 0.0


def _to_int(value: Any) -> int:
    return int(round(_to_float(value)))


def _detect_column(headers: list[str], aliases: list[str]) -> int:
    normalized_headers = [_normalize(h) for h in headers]
    normalized_aliases = [_normalize(a) for a in aliases]
    for alias in normalized_aliases:
        for idx, header in enumerate(normalized_headers):
            if header == alias:
                return idx
    for alias in normalized_aliases:
        for idx, header in enumerate(normalized_headers):
            if alias and alias in header:
                return idx
    return -1


def _iter_sheet_rows(file_path: Path) -> list[list[Any]]:
    if file_path.suffix.lower() == '.csv':
        with file_path.open('r', encoding='utf-8-sig', newline='') as handle:
            return [row for row in csv.reader(handle)]

    workbook = load_workbook(file_path, data_only=True)
    try:
        sheet = workbook.active
        rows: list[list[Any]] = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(list(row))
        return rows
    finally:
        workbook.close()


def _cache_path_for_file(file_path: Path, range_value: str, cache_root: Path) -> Path:
    stats = file_path.stat()
    fingerprint = '|'.join(
        [
            'v2',
            str(file_path),
            str(stats.st_size),
            str(int(stats.st_mtime)),
            range_value,
        ]
    )
    return cache_root / f'{hashlib.md5(fingerprint.encode("utf-8")).hexdigest()}.json'


def _term_type(term: str) -> str:
    clean = (term or '').strip().upper()
    return 'asin' if clean.startswith('B0') else 'keyword'


def _filter_by_term_type(rows: list[dict[str, Any]], term_type_filter: str) -> list[dict[str, Any]]:
    if term_type_filter == 'all':
        return rows
    return [row for row in rows if _term_type(str(row.get('search_term', ''))) == term_type_filter]


def _text_contains(haystack: str, needle: str) -> bool:
    return needle.strip().lower() in haystack.strip().lower()


def _sort_desc(rows: list[dict[str, Any]], key: str) -> list[dict[str, Any]]:
    return sorted(rows, key=lambda row: float(row.get(key, 0.0) or 0.0), reverse=True)


def _sort_asc(rows: list[dict[str, Any]], key: str) -> list[dict[str, Any]]:
    return sorted(rows, key=lambda row: float(row.get(key, 0.0) or 0.0))


def build_report_from_file(file_path: Path, range_value: str) -> dict[str, Any]:
    rows = _iter_sheet_rows(file_path)
    if not rows:
        raise ValueError('No rows found in search term file.')

    header_idx = -1
    for idx, row in enumerate(rows[:100]):
        lowered = ' '.join(str(cell or '') for cell in row).lower()
        if 'search term' in lowered and 'click' in lowered and 'spend' in lowered:
            header_idx = idx
            break
    if header_idx < 0:
        raise ValueError('Search term headers were not found.')

    header_row = [str(cell or '').strip() for cell in rows[header_idx]]

    col_term = _detect_column(header_row, ['Customer Search Term', 'Search Term', 'Customer Search Query'])
    col_campaign = _detect_column(header_row, ['Campaign Name'])
    col_ad_group = _detect_column(header_row, ['Ad Group Name'])
    col_match = _detect_column(header_row, ['Match Type'])
    col_impressions = _detect_column(header_row, ['Impressions'])
    col_clicks = _detect_column(header_row, ['Clicks'])
    col_spend = _detect_column(header_row, ['Spend'])
    col_sales = _detect_column(header_row, ['7 Day Total Sales', '14 Day Total Sales', 'Total Sales', 'Attributed Sales'])
    col_orders = _detect_column(header_row, ['7 Day Total Orders', '14 Day Total Orders', 'Total Orders', 'Attributed Conversions'])

    if col_term < 0 or col_clicks < 0 or col_spend < 0:
        raise ValueError('Required search term columns were not found.')

    grouped: dict[str, dict[str, Any]] = {}
    for row in rows[header_idx + 1 :]:
        if not row:
            continue
        term = str(row[col_term] if col_term < len(row) else '').strip()
        if term == '' or term.lower() in {'nan', 'none'}:
            continue

        item = grouped.setdefault(
            term.lower(),
            {
                'search_term': term,
                'campaign_name': '',
                'ad_group_name': '',
                'match_type': '',
                'impressions': 0.0,
                'clicks': 0.0,
                'spend': 0.0,
                'sales': 0.0,
                'orders': 0.0,
            },
        )

        if item['campaign_name'] == '' and col_campaign >= 0 and col_campaign < len(row):
            item['campaign_name'] = str(row[col_campaign] or '').strip()
        if item['ad_group_name'] == '' and col_ad_group >= 0 and col_ad_group < len(row):
            item['ad_group_name'] = str(row[col_ad_group] or '').strip()
        if item['match_type'] == '' and col_match >= 0 and col_match < len(row):
            item['match_type'] = str(row[col_match] or '').strip()

        item['impressions'] += _to_float(row[col_impressions]) if col_impressions >= 0 and col_impressions < len(row) else 0.0
        item['clicks'] += _to_float(row[col_clicks]) if col_clicks >= 0 and col_clicks < len(row) else 0.0
        item['spend'] += _to_float(row[col_spend]) if col_spend >= 0 and col_spend < len(row) else 0.0
        item['sales'] += _to_float(row[col_sales]) if col_sales >= 0 and col_sales < len(row) else 0.0
        item['orders'] += _to_float(row[col_orders]) if col_orders >= 0 and col_orders < len(row) else 0.0

    if not grouped:
        raise ValueError('No valid search term rows were found.')

    all_terms: list[dict[str, Any]] = []
    for item in grouped.values():
        clicks = float(item['clicks'])
        impressions = float(item['impressions'])
        spend = float(item['spend'])
        sales = float(item['sales'])
        orders = float(item['orders'])
        all_terms.append(
            {
                'search_term': str(item['search_term']),
                'campaign_name': str(item['campaign_name']),
                'ad_group_name': str(item['ad_group_name']),
                'match_type': str(item['match_type']),
                'impressions': _to_int(impressions),
                'clicks': _to_int(clicks),
                'spend': round(spend, 2),
                'sales': round(sales, 2),
                'orders': _to_int(orders),
                'ctr': (clicks / impressions * 100.0) if impressions > 0 else 0.0,
                'cpc': (spend / clicks) if clicks > 0 else 0.0,
                'cvr': (orders / clicks * 100.0) if clicks > 0 else 0.0,
                'acos': (spend / sales * 100.0) if sales > 0 else 0.0,
                'roas': (sales / spend) if spend > 0 else 0.0,
            }
        )

    summary = {
        'rows': len(all_terms),
        'terms': len(all_terms),
        'impressions': int(sum(float(r['impressions']) for r in all_terms)),
        'clicks': int(sum(float(r['clicks']) for r in all_terms)),
        'spend': float(sum(float(r['spend']) for r in all_terms)),
        'sales': float(sum(float(r['sales']) for r in all_terms)),
        'orders': int(sum(float(r['orders']) for r in all_terms)),
    }
    summary['ctr'] = (summary['clicks'] / summary['impressions'] * 100.0) if summary['impressions'] > 0 else 0.0
    summary['cpc'] = (summary['spend'] / summary['clicks']) if summary['clicks'] > 0 else 0.0
    summary['cvr'] = (summary['orders'] / summary['clicks'] * 100.0) if summary['clicks'] > 0 else 0.0
    summary['acos'] = (summary['spend'] / summary['sales'] * 100.0) if summary['sales'] > 0 else 0.0
    summary['roas'] = (summary['sales'] / summary['spend']) if summary['spend'] > 0 else 0.0
    summary['wasted_spend'] = float(sum(float(r['spend']) for r in all_terms if float(r['sales']) <= 0))
    summary['wasted_spend_pct'] = (summary['wasted_spend'] / summary['spend'] * 100.0) if summary['spend'] > 0 else 0.0

    thresholds = {'spend': 800, 'clicks': 24}

    top_by_sales = _sort_desc(all_terms, 'sales')[:25]
    top_by_spend = _sort_desc(all_terms, 'spend')[:25]
    winners = _sort_desc([r for r in all_terms if r['orders'] >= 2 and r['sales'] > 0 and r['acos'] <= 30], 'orders')[:25]
    high_acos = _sort_desc([r for r in all_terms if r['sales'] > 0 and r['acos'] >= 30], 'acos')[:25]
    high_roas = _sort_desc([r for r in all_terms if r['sales'] > 0], 'roas')[:25]
    low_roas = _sort_asc([r for r in all_terms if r['sales'] > 0], 'roas')[:25]
    high_spend_no_sales = _sort_desc([r for r in all_terms if r['orders'] <= 0 and r['spend'] >= thresholds['spend']], 'spend')[:25]
    high_clicks_no_orders = _sort_desc([r for r in all_terms if r['orders'] <= 0 and r['clicks'] >= thresholds['clicks']], 'clicks')[:25]
    top_ctr = _sort_desc([r for r in all_terms if r['impressions'] >= 200], 'ctr')[:25]
    low_ctr = _sort_asc([r for r in all_terms if r['impressions'] >= 200], 'ctr')[:25]
    hidden_gems = _sort_desc([r for r in all_terms if r['orders'] > 0 and r['clicks'] <= 12 and r['roas'] >= 2.5], 'roas')[:25]

    report = {
        'summary': summary,
        'thresholds': thresholds,
        'all_terms': _sort_desc(all_terms, 'spend'),
        'detail_rows': _sort_desc(all_terms, 'spend'),
        'top_by_sales': top_by_sales,
        'top_by_spend': top_by_spend,
        'high_acos': high_acos,
        'high_roas': high_roas,
        'low_roas': low_roas,
        'winners': winners,
        'high_spend_no_sales': high_spend_no_sales,
        'high_clicks_no_orders': high_clicks_no_orders,
        'top_ctr': top_ctr,
        'low_ctr': low_ctr,
        'hidden_gems': hidden_gems,
    }
    return report


def load_or_build_report(
    *,
    file_path: Path,
    range_value: str,
    cache_dir_candidates: list[Path],
) -> dict[str, Any]:
    for cache_root in cache_dir_candidates:
        cache_path = _cache_path_for_file(file_path, range_value, cache_root)
        if cache_path.exists():
            try:
                data = json.loads(cache_path.read_text(encoding='utf-8'))
                if isinstance(data, dict) and isinstance(data.get('summary'), dict):
                    return data
            except (OSError, json.JSONDecodeError):
                pass

    report = build_report_from_file(file_path, range_value)

    for cache_root in cache_dir_candidates:
        try:
            cache_root.mkdir(parents=True, exist_ok=True)
            cache_path = _cache_path_for_file(file_path, range_value, cache_root)
            cache_path.write_text(json.dumps(report, ensure_ascii=False), encoding='utf-8')
            break
        except OSError:
            continue

    return report


def build_section_rows(report: dict[str, Any], term_type_filter: str, rows_limit: int) -> dict[str, list[dict[str, Any]]]:
    def _take(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        filtered = _filter_by_term_type(rows, term_type_filter)
        if rows_limit <= 0:
            return filtered
        return filtered[:rows_limit]

    return {
        'top_by_sales': _take(list(report.get('top_by_sales', []))),
        'top_by_spend': _take(list(report.get('top_by_spend', []))),
        'winners': _take(list(report.get('winners', []))),
        'high_acos': _take(list(report.get('high_acos', []))),
        'high_roas': _take(list(report.get('high_roas', []))),
        'low_roas': _take(list(report.get('low_roas', []))),
        'high_spend_no_sales': _take(list(report.get('high_spend_no_sales', []))),
        'high_clicks_no_orders': _take(list(report.get('high_clicks_no_orders', []))),
        'top_ctr': _take(list(report.get('top_ctr', []))),
        'low_ctr': _take(list(report.get('low_ctr', []))),
        'hidden_gems': _take(list(report.get('hidden_gems', []))),
    }


def build_plan_rows(
    report: dict[str, Any],
    targeting_query: str,
    term_type_filter: str,
    rows_limit: int,
) -> list[dict[str, Any]]:
    query = (targeting_query or '').strip()
    if query == '':
        return []

    base_rows = list(report.get('detail_rows', []))
    filtered = _filter_by_term_type(base_rows, term_type_filter)

    out: list[dict[str, Any]] = []
    for row in filtered:
        search_term = str(row.get('search_term', ''))
        campaign = str(row.get('campaign_name', ''))
        ad_group = str(row.get('ad_group_name', ''))
        match = _text_contains(search_term, query) or _text_contains(campaign, query) or _text_contains(ad_group, query)
        clicks = int(row.get('clicks', 0) or 0)
        orders = int(row.get('orders', 0) or 0)
        sales = float(row.get('sales', 0.0) or 0.0)
        spend = float(row.get('spend', 0.0) or 0.0)
        roas = float(row.get('roas', 0.0) or 0.0)
        acos = float(row.get('acos', 0.0) or 0.0)

        if match and sales > 0 and roas >= 2:
            action = 'Keep and scale bid carefully.'
        elif match and clicks >= 20 and orders == 0:
            action = 'Lower bid or add negative variant.'
        elif not match and spend > 0:
            action = 'Check relevance. Consider negative targeting.'
        else:
            action = 'Monitor and optimize based on next cycle.'

        out.append(
            {
                'targeting_text': query,
                'customer_search_term': search_term,
                'match_type': str(row.get('match_type', '')),
                'match_result': 'Match' if match else 'Not Match',
                'campaign_name': campaign,
                'ad_group_name': ad_group,
                'spend': spend,
                'sales': sales,
                'orders': orders,
                'acos': acos,
                'roas': roas,
                'plan_action': action,
            }
        )

    if rows_limit > 0:
        out = out[:rows_limit]
    return out
