from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from ..db import transaction, upsert_app_setting


METRIC_MAP = {
    'impression': 'impressions',
    'impressions': 'impressions',
    'click': 'clicks',
    'clicks': 'clicks',
    'page view': 'page_views',
    'page views': 'page_views',
    'session': 'sessions',
    'sessions': 'sessions',
    'ctr': 'ctr',
    'ads spend': 'spend',
    'spend': 'spend',
    'ads sales': 'sales',
    'sales': 'sales',
    'total unit': 'total_units',
    'total units': 'total_units',
    'total sale': 'total_sales',
    'total sales': 'total_sales',
    'acos': 'acos',
    'tacos': 'tacos',
    'conversion': 'conversion_rate',
    'conversion rate': 'conversion_rate',
}


def clean_number(value: Any) -> float:
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    raw = str(value).strip()
    if raw == '':
        return 0.0

    is_negative = False
    if raw.startswith('(') and raw.endswith(')'):
        is_negative = True
        raw = raw[1:-1]

    # Keep only numeric separators/sign; remove currency symbols/text.
    txt = re.sub(r'[^0-9,.\-]+', '', raw)
    if txt in ('', '-', '.', '-.', ',', '-,'):
        return 0.0

    # Normalize decimal/thousand separators for common locales.
    if ',' in txt and '.' in txt:
        # Treat rightmost separator as decimal point.
        if txt.rfind(',') > txt.rfind('.'):
            txt = txt.replace('.', '')
            txt = txt.replace(',', '.')
        else:
            txt = txt.replace(',', '')
    elif ',' in txt and '.' not in txt:
        # Single comma may be decimal; multiple are likely thousand separators.
        if txt.count(',') == 1 and len(txt.split(',')[-1]) in (1, 2):
            txt = txt.replace(',', '.')
        else:
            txt = txt.replace(',', '')

    if is_negative and not txt.startswith('-'):
        txt = '-' + txt

    try:
        return float(txt)
    except ValueError:
        return 0.0


def normalize_header(value: Any) -> str:
    txt = re.sub(r'[^a-z0-9]+', ' ', str(value or '').strip().lower())
    return re.sub(r'\s+', ' ', txt).strip()


def map_metric_header(header: str) -> str | None:
    return METRIC_MAP.get(header)


def normalize_month_name(value: Any) -> str | None:
    if value is None or value == '':
        return None

    if isinstance(value, datetime):
        return value.strftime('%b-%y')

    if isinstance(value, (int, float)):
        num = float(value)
        if 30000 <= num <= 60000:
            try:
                return from_excel(num).strftime('%b-%y')
            except Exception:
                return None

    text = str(value).strip()
    if not text:
        return None

    match = re.search(r'\b(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\b[^0-9]*(\d{2,4})', text, re.I)
    if match:
        month = match.group(1)[:3].title()
        year = int(match.group(2)) % 100
        return f'{month}-{year:02d}'

    return None


def get_cell_value(ws, row: int, col: int):
    return ws.cell(row=row, column=col).value


def detect_header_row(ws) -> int | None:
    max_col = ws.max_column
    for row in range(1, min(10, ws.max_row) + 1):
        has_sku = False
        has_asin = False
        for col in range(1, max_col + 1):
            header = normalize_header(get_cell_value(ws, row, col))
            if header in ('sku', 'skus'):
                has_sku = True
            if 'asin' in header:
                has_asin = True
        if has_sku and has_asin:
            return row
    return None


def find_source_sheet(workbook):
    mom_sheet = None
    for ws in workbook.worksheets:
        if ws.title.strip().lower() == 'mom':
            mom_sheet = ws
            break
    if mom_sheet and detect_header_row(mom_sheet) is not None:
        return mom_sheet

    for ws in workbook.worksheets:
        if detect_header_row(ws) is not None:
            return ws

    raise RuntimeError('A valid upload sheet with SKU and ASIN headers could not be found.')


def get_identifier_columns(ws, header_row: int) -> dict[str, int]:
    cols: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = normalize_header(get_cell_value(ws, header_row, col))
        if 'sku' in (' ' + header + ' ') and 'sku' not in cols and header in ('sku', 'skus'):
            cols['sku'] = col
        elif 'asin' in header and 'asin' not in cols:
            cols['asin'] = col
        elif header in ('category', 'categories') and 'category' not in cols:
            cols['category'] = col

    if 'sku' not in cols or 'asin' not in cols:
        raise RuntimeError('The upload sheet must contain SKU and ASIN columns.')

    return cols


def get_month_label(ws, row: int, col: int) -> str | None:
    return normalize_month_name(get_cell_value(ws, row, col))


def detect_month_row(ws, header_row: int) -> int:
    best_row = max(1, header_row - 2)
    best_count = 0

    for row in range(1, header_row):
        count = 0
        for col in range(1, ws.max_column + 1):
            metric = map_metric_header(normalize_header(get_cell_value(ws, header_row, col)))
            if metric in ('sessions', 'impressions') and get_month_label(ws, row, col) is not None:
                count += 1
        if count > best_count:
            best_count = count
            best_row = row

    if best_count == 0:
        raise RuntimeError('Month headers could not be detected in the uploaded file.')

    return best_row


def build_month_configs(ws, month_row: int, header_row: int) -> list[dict[str, Any]]:
    month_configs: dict[str, dict[str, Any]] = {}
    current_month: str | None = None

    for col in range(1, ws.max_column + 1):
        month_candidate = get_month_label(ws, month_row, col)
        if month_candidate is not None:
            current_month = month_candidate
            if current_month not in month_configs:
                month_configs[current_month] = {'month': current_month, 'columns': {}}

        if current_month is None:
            continue

        metric = map_metric_header(normalize_header(get_cell_value(ws, header_row, col)))
        if metric and metric not in month_configs[current_month]['columns']:
            month_configs[current_month]['columns'][metric] = col

    configs = [
        cfg
        for cfg in month_configs.values()
        if 'spend' in cfg['columns'] and 'sales' in cfg['columns'] and len(cfg['columns']) >= 5
    ]

    if not configs:
        raise RuntimeError('No month blocks were detected in the uploaded file.')

    return configs


def percentage(numerator: float, denominator: float) -> float:
    if denominator <= 0:
        return 0.0
    return (numerator / denominator) * 100


def get_metric_number(ws, row: int, columns: dict[str, int], metric: str) -> float:
    col = columns.get(metric)
    if not col:
        return 0.0
    return clean_number(get_cell_value(ws, row, col))


def summarize_months(month_configs: list[dict[str, Any]]) -> str:
    months = [cfg['month'] for cfg in month_configs]
    if not months:
        return '0 months'
    if len(months) == 1:
        return f"1 month ({months[0]})"
    return f"{len(months)} months ({months[0]} to {months[-1]})"


def _replace_monthly_ads(
    rows: list[dict[str, Any]],
    *,
    app_settings: dict[str, str | None] | None = None,
) -> None:
    insert_sql = """
        INSERT INTO monthly_ads (
            user_id, sku, asin, category, month_name,
            impressions, clicks, page_views, sessions, ctr,
            spend, sales, total_units, total_sales, acos, tacos, conversion_rate
        ) VALUES (
            %s, %s, %s, %s, %s,
            %s, %s, %s, %s, %s,
            %s, %s, %s, %s, %s, %s, %s
        )
    """
    with transaction() as connection:
        with connection.cursor() as cursor:
            user_ids = sorted({row.get('user_id') for row in rows if row.get('user_id') is not None})
            if user_ids:
                cursor.execute("DELETE FROM monthly_ads WHERE user_id = ANY(%s)", (user_ids,))
            else:
                cursor.execute('DELETE FROM monthly_ads')
            if rows:
                cursor.executemany(
                    insert_sql,
                    [
                        (
                            row.get('user_id'),
                            row.get('sku'),
                            row.get('asin'),
                            row.get('category'),
                            row.get('month_name'),
                            row.get('impressions'),
                            row.get('clicks'),
                            row.get('page_views'),
                            row.get('sessions'),
                            row.get('ctr'),
                            row.get('spend'),
                            row.get('sales'),
                            row.get('total_units'),
                            row.get('total_sales'),
                            row.get('acos'),
                            row.get('tacos'),
                            row.get('conversion_rate'),
                        )
                        for row in rows
                    ],
                )
            if app_settings:
                for key, value in app_settings.items():
                    upsert_app_setting(key, value, connection=connection)


def _normalize_col_name(value: Any) -> str:
    return normalize_header(value)


def _find_column(
    columns: list[str],
    includes: list[str],
    *,
    excludes: list[str] | None = None,
) -> str | None:
    excludes = excludes or []
    for col in columns:
        key = _normalize_col_name(col)
        if all(token in key for token in includes) and not any(token in key for token in excludes):
            return col
    return None


def _normalize_asin(value: Any) -> str:
    txt = str(value or '').strip().upper()
    txt = re.sub(r'[^A-Z0-9]+', '', txt)
    if txt in ('', 'NAN', 'NONE'):
        return ''
    return txt


def _split_creative_asins(value: Any) -> list[str]:
    raw = str(value or '').strip()
    if raw == '':
        return []
    tokens = re.split(r'[,;\n\r]+', raw)
    asins: list[str] = []
    for token in tokens:
        asin = _normalize_asin(token)
        if asin != '':
            asins.append(asin)
    # preserve order, remove duplicates
    seen: set[str] = set()
    deduped: list[str] = []
    for asin in asins:
        if asin in seen:
            continue
        seen.add(asin)
        deduped.append(asin)
    return deduped


def _sheet_for_type(sheet_names: list[str], marker: str) -> str | None:
    marker_norm = normalize_header(marker)
    for name in sheet_names:
        if marker_norm in normalize_header(name):
            return name
    return None


def _parse_bulk_ads_metrics(file_path: Path) -> tuple[dict[str, dict[str, float]], dict[str, int], dict[str, dict[str, float]]]:
    excel = pd.ExcelFile(file_path)
    sheet_names = list(excel.sheet_names)
    result: dict[str, dict[str, float]] = {}
    source_rows = {'sp': 0, 'sd': 0, 'sb': 0}
    type_totals = {
        'SP': {'impressions': 0.0, 'clicks': 0.0, 'spend': 0.0, 'sales': 0.0},
        'SD': {'impressions': 0.0, 'clicks': 0.0, 'spend': 0.0, 'sales': 0.0},
        'SB': {'impressions': 0.0, 'clicks': 0.0, 'spend': 0.0, 'sales': 0.0},
        'SBV': {'impressions': 0.0, 'clicks': 0.0, 'spend': 0.0, 'sales': 0.0},
    }

    def add_metrics(asin: str, impressions: float, clicks: float, spend: float, sales: float) -> None:
        if asin == '':
            return
        bucket = result.setdefault(asin, {'impressions': 0.0, 'clicks': 0.0, 'spend': 0.0, 'sales': 0.0})
        bucket['impressions'] += impressions
        bucket['clicks'] += clicks
        bucket['spend'] += spend
        bucket['sales'] += sales

    def parse_sp_or_sd(sheet_name: str, source_key: str, ad_type: str) -> None:
        frame = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str).fillna('')
        columns = list(frame.columns)
        asin_col = _find_column(columns, ['asin', 'informational', 'only']) or _find_column(columns, ['asin'])
        impressions_col = _find_column(columns, ['impressions'])
        clicks_col = _find_column(columns, ['clicks'])
        spend_col = _find_column(columns, ['spend']) or _find_column(columns, ['cost'])
        sales_col = _find_column(columns, ['sales'])

        if not asin_col:
            raise RuntimeError(f'ASIN column not found in sheet "{sheet_name}".')

        for _, row in frame.iterrows():
            asin = _normalize_asin(row.get(asin_col, ''))
            if asin == '':
                continue
            source_rows[source_key] += 1
            add_metrics(
                asin,
                clean_number(row.get(impressions_col, '')) if impressions_col else 0.0,
                clean_number(row.get(clicks_col, '')) if clicks_col else 0.0,
                clean_number(row.get(spend_col, '')) if spend_col else 0.0,
                clean_number(row.get(sales_col, '')) if sales_col else 0.0,
            )
            type_totals[ad_type]['impressions'] += clean_number(row.get(impressions_col, '')) if impressions_col else 0.0
            type_totals[ad_type]['clicks'] += clean_number(row.get(clicks_col, '')) if clicks_col else 0.0
            type_totals[ad_type]['spend'] += clean_number(row.get(spend_col, '')) if spend_col else 0.0
            type_totals[ad_type]['sales'] += clean_number(row.get(sales_col, '')) if sales_col else 0.0

    def parse_sb(sheet_name: str) -> None:
        frame = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str).fillna('')
        columns = list(frame.columns)
        creative_asins_col = _find_column(columns, ['creative', 'asins']) or _find_column(columns, ['creative', 'asin'])
        campaign_name_col = _find_column(columns, ['campaign', 'name']) or _find_column(columns, ['campaign'])
        impressions_col = _find_column(columns, ['impressions'])
        clicks_col = _find_column(columns, ['clicks'])
        spend_col = _find_column(columns, ['spend']) or _find_column(columns, ['cost'])
        sales_col = _find_column(columns, ['sales'])

        if not creative_asins_col:
            raise RuntimeError(f'Creative ASINs column not found in sheet "{sheet_name}".')

        for _, row in frame.iterrows():
            asins = _split_creative_asins(row.get(creative_asins_col, ''))
            if not asins:
                continue

            source_rows['sb'] += 1
            divisor = float(len(asins))
            impressions_each = (clean_number(row.get(impressions_col, '')) if impressions_col else 0.0) / divisor
            clicks_each = (clean_number(row.get(clicks_col, '')) if clicks_col else 0.0) / divisor
            spend_each = (clean_number(row.get(spend_col, '')) if spend_col else 0.0) / divisor
            sales_each = (clean_number(row.get(sales_col, '')) if sales_col else 0.0) / divisor
            campaign_name = str(row.get(campaign_name_col, '') or '') if campaign_name_col else ''
            campaign_name_upper = campaign_name.strip().upper()
            sb_bucket = 'SBV' if ('SBV' in campaign_name_upper or 'VIDEO' in campaign_name_upper) else 'SB'

            for asin in asins:
                add_metrics(asin, impressions_each, clicks_each, spend_each, sales_each)
            type_totals[sb_bucket]['impressions'] += impressions_each * divisor
            type_totals[sb_bucket]['clicks'] += clicks_each * divisor
            type_totals[sb_bucket]['spend'] += spend_each * divisor
            type_totals[sb_bucket]['sales'] += sales_each * divisor

    sp_sheet = _sheet_for_type(sheet_names, 'Sponsored Product')
    sd_sheet = _sheet_for_type(sheet_names, 'Sponsored Display')
    sb_sheet = _sheet_for_type(sheet_names, 'Sponsored Brand')

    if not any([sp_sheet, sd_sheet, sb_sheet]):
        raise RuntimeError('Sponsored Product / Sponsored Display / Sponsored Brand sheets were not found in bulk file.')

    if sp_sheet:
        parse_sp_or_sd(sp_sheet, 'sp', 'SP')
    if sd_sheet:
        parse_sp_or_sd(sd_sheet, 'sd', 'SD')
    if sb_sheet:
        parse_sb(sb_sheet)

    return result, source_rows, type_totals


def _load_business_metrics(file_path: Path) -> dict[str, dict[str, float]]:
    suffix = file_path.suffix.lower()
    if suffix == '.csv':
        frame = pd.read_csv(file_path, dtype=str).fillna('')
    else:
        frame = pd.read_excel(file_path, dtype=str).fillna('')

    columns = list(frame.columns)
    asin_col = _find_column(columns, ['asin'])
    units_col = _find_column(columns, ['units', 'ordered'], excludes=['b2b'])
    ordered_sales_col = _find_column(columns, ['ordered', 'product', 'sales'], excludes=['b2b'])
    sessions_col = _find_column(columns, ['sessions', 'total'], excludes=['b2b']) or _find_column(columns, ['session'])
    page_views_col = (
        _find_column(columns, ['page', 'views', 'total'], excludes=['b2b'])
        or _find_column(columns, ['page', 'view', 'total'], excludes=['b2b'])
        or _find_column(columns, ['page', 'views'])
        or _find_column(columns, ['page', 'view'])
    )

    if not asin_col:
        raise RuntimeError('ASIN column not found in Business Report file.')
    if not units_col:
        raise RuntimeError('Units Ordered column not found in Business Report file.')
    if not ordered_sales_col:
        raise RuntimeError('Ordered Product Sales column not found in Business Report file.')

    result: dict[str, dict[str, float]] = {}
    for _, row in frame.iterrows():
        asin = _normalize_asin(row.get(asin_col, ''))
        if asin == '':
            continue
        bucket = result.setdefault(asin, {'total_units': 0.0, 'total_sales': 0.0, 'sessions': 0.0, 'page_views': 0.0})
        bucket['total_units'] += clean_number(row.get(units_col, ''))
        bucket['total_sales'] += clean_number(row.get(ordered_sales_col, ''))
        if sessions_col:
            bucket['sessions'] += clean_number(row.get(sessions_col, ''))
        if page_views_col:
            bucket['page_views'] += clean_number(row.get(page_views_col, ''))
    return result


def _extract_month_label_from_filename(file_path: Path) -> str:
    name = file_path.name
    match = re.search(r'(\d{8})-(\d{8})', name)
    if match:
        end_raw = match.group(2)
        try:
            dt = datetime.strptime(end_raw, '%Y%m%d')
            return dt.strftime('%b-%y')
        except ValueError:
            pass
    return datetime.utcnow().strftime('%b-%y')


def parse_monthly_ads_excel(file_path: Path, business_report_path: Path | None = None, user_id: int | None = None) -> tuple[int, str]:
    if business_report_path is not None:
        try:
            ads_metrics, source_rows, type_totals = _parse_bulk_ads_metrics(file_path)
            business_metrics = _load_business_metrics(business_report_path)

            all_asins = sorted(set(ads_metrics.keys()) | set(business_metrics.keys()))
            if not all_asins:
                return 0, 'No valid ASIN rows found after processing bulk and business files.'

            month_label = _extract_month_label_from_filename(file_path)
            to_insert: list[dict[str, Any]] = []

            for asin in all_asins:
                ads_row = ads_metrics.get(asin, {})
                business_row = business_metrics.get(asin, {})

                impressions = int(round(float(ads_row.get('impressions', 0.0))))
                clicks = int(round(float(ads_row.get('clicks', 0.0))))
                spend = float(ads_row.get('spend', 0.0))
                sales = float(ads_row.get('sales', 0.0))
                total_units = int(round(float(business_row.get('total_units', 0.0))))
                total_sales = float(business_row.get('total_sales', 0.0))
                sessions = int(round(float(business_row.get('sessions', 0.0))))
                page_views = int(round(float(business_row.get('page_views', 0.0))))

                ctr = percentage(clicks, impressions)
                acos = percentage(spend, sales)
                tacos = percentage(spend, total_sales)
                conversion_denominator = sessions if sessions > 0 else clicks
                conversion_rate = percentage(total_units, conversion_denominator)

                to_insert.append(
                    {
                        'user_id': user_id,
                        'sku': asin,
                        'asin': asin,
                        'category': None,
                        'month_name': month_label,
                        'impressions': impressions,
                        'clicks': clicks,
                        'page_views': page_views,
                        'sessions': sessions,
                        'ctr': ctr,
                        'spend': spend,
                        'sales': sales,
                        'total_units': total_units,
                        'total_sales': total_sales,
                        'acos': acos,
                        'tacos': tacos,
                        'conversion_rate': conversion_rate,
                    }
                )

            _replace_monthly_ads(
                to_insert,
                app_settings={
                    (f'mom_ad_type_share_user_{user_id}' if user_id is not None else 'mom_ad_type_share'): json.dumps(
                        type_totals,
                        separators=(',', ':'),
                    ),
                },
            )

            message = f'Upload completed successfully for {month_label}.'
            return len(to_insert), message
        except Exception as exc:
            return 0, str(exc)

    workbook = None
    try:
        workbook = load_workbook(filename=file_path, data_only=True)
        ws = find_source_sheet(workbook)
        header_row = detect_header_row(ws)
        if header_row is None:
            return 0, 'The upload sheet headers could not be recognized.'

        month_row = detect_month_row(ws, header_row)
        identifier_columns = get_identifier_columns(ws, header_row)
        month_configs = build_month_configs(ws, month_row, header_row)

        source_rows = 0
        to_insert: list[dict[str, Any]] = []

        for row in range(header_row + 1, ws.max_row + 1):
            sku = str(get_cell_value(ws, row, identifier_columns['sku']) or '').strip()
            asin = str(get_cell_value(ws, row, identifier_columns['asin']) or '').strip()
            category = str(get_cell_value(ws, row, identifier_columns['category']) or '').strip() if 'category' in identifier_columns else ''

            if not sku and not asin:
                continue

            source_rows += 1

            for cfg in month_configs:
                cols = cfg['columns']
                impressions = int(get_metric_number(ws, row, cols, 'impressions'))
                clicks = int(get_metric_number(ws, row, cols, 'clicks'))
                page_views = int(get_metric_number(ws, row, cols, 'page_views'))
                sessions = int(get_metric_number(ws, row, cols, 'sessions'))
                spend = get_metric_number(ws, row, cols, 'spend')
                sales = get_metric_number(ws, row, cols, 'sales')
                total_units = int(get_metric_number(ws, row, cols, 'total_units'))
                total_sales = get_metric_number(ws, row, cols, 'total_sales')

                ctr = get_metric_number(ws, row, cols, 'ctr') * 100 if 'ctr' in cols else percentage(clicks, impressions)
                acos = get_metric_number(ws, row, cols, 'acos') * 100 if 'acos' in cols else percentage(spend, sales)
                tacos = get_metric_number(ws, row, cols, 'tacos') * 100 if 'tacos' in cols else percentage(spend, total_sales)
                conversion_rate = get_metric_number(ws, row, cols, 'conversion_rate') * 100 if 'conversion_rate' in cols else percentage(total_units, sessions)

                to_insert.append(
                    {
                        'user_id': user_id,
                        'sku': sku or None,
                        'asin': asin or None,
                        'category': category or None,
                        'month_name': cfg['month'],
                        'impressions': impressions,
                        'clicks': clicks,
                        'page_views': page_views,
                        'sessions': sessions,
                        'ctr': ctr,
                        'spend': spend,
                        'sales': sales,
                        'total_units': total_units,
                        'total_sales': total_sales,
                        'acos': acos,
                        'tacos': tacos,
                        'conversion_rate': conversion_rate,
                    }
                )

        if not to_insert:
            return 0, 'No valid data rows were found in the uploaded file.'

        # Same behavior as PHP app: replace previous monthly data with newly uploaded file.
        _replace_monthly_ads(to_insert)

        message = f"Upload completed successfully for {summarize_months(month_configs)}."
        return len(to_insert), message
    except Exception as exc:
        return 0, str(exc)
    finally:
        if workbook is not None:
            workbook.close()
