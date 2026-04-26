from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..db import transaction


def _normalize_text(value: Any) -> str:
    txt = re.sub(r'[^a-z0-9]+', ' ', str(value or '').strip().lower())
    return re.sub(r'\s+', ' ', txt).strip()


def _to_number(value: Any) -> float:
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    txt = re.sub(r'[^0-9.\-]+', '', str(value))
    if txt in ('', '-', '.', '-.'):
        return 0.0
    try:
        return float(txt)
    except ValueError:
        return 0.0


def _cell(ws, row: int, col: int) -> Any:
    return ws.cell(row=row, column=col).value


def _extract_week_label(value: Any) -> str | None:
    txt = str(value or '').strip()
    if not txt:
        return None
    pattern = re.compile(
        r'\b(\d{1,2}\s*[A-Za-z]{3}(?:[-\']?\d{2,4})?)\s*-\s*(\d{1,2}\s*[A-Za-z]{3}(?:[-\']?\d{2,4})?)\b',
        re.I,
    )
    m = pattern.search(txt)
    if not m:
        return None
    left = re.sub(r'\s+', ' ', m.group(1).strip())
    right = re.sub(r'\s+', ' ', m.group(2).strip())
    return f'{left} - {right}'


def _detect_week_blocks(ws) -> list[dict[str, int | str]]:
    highest_row = min(ws.max_row, 30)
    blocks: list[dict[str, int | str]] = []
    for row in range(1, highest_row + 1):
        for col in range(1, ws.max_column + 1):
            label = _extract_week_label(_cell(ws, row, col))
            if label:
                blocks.append({'week_range': label, 'header_anchor_row': row, 'start_col': col})

    if not blocks:
        return []

    blocks.sort(key=lambda b: (int(b['header_anchor_row']), int(b['start_col'])))
    deduped: list[dict[str, int | str]] = []
    seen: set[str] = set()
    for block in blocks:
        key = f"{block['header_anchor_row']}|{block['start_col']}"
        if key in seen:
            continue
        seen.add(key)
        deduped.append(block)

    for i, block in enumerate(deduped):
        next_start = int(deduped[i + 1]['start_col']) if i + 1 < len(deduped) else (ws.max_column + 1)
        block['end_col'] = max(int(block['start_col']), next_start - 1)

    return deduped


def _metric_map_for_row(ws, row: int, start_col: int, end_col: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    fallback_sales: int | None = None
    for col in range(start_col, end_col + 1):
        header = _normalize_text(_cell(ws, row, col))
        if not header:
            continue
        if 'impression' in header and 'impressions' not in mapping:
            mapping['impressions'] = col
        elif 'click' in header and 'clicks' not in mapping:
            mapping['clicks'] = col
        elif ('page view' in header or 'page views' in header) and 'page_views' not in mapping:
            mapping['page_views'] = col
        elif 'session' in header and 'sessions' not in mapping:
            mapping['sessions'] = col
        elif ('ads spend' in header or header == 'spend') and 'spend' not in mapping:
            mapping['spend'] = col
        elif 'sales' in header and 'sales' not in mapping:
            if 'total sales' in header:
                fallback_sales = col
            else:
                mapping['sales'] = col
        elif ('total unit' in header or 'units' in header) and 'total_units' not in mapping:
            mapping['total_units'] = col
        elif 'total sales' in header and 'total_sales' not in mapping:
            mapping['total_sales'] = col
        elif header == 'ctr' and 'ctr' not in mapping:
            mapping['ctr'] = col
        elif header == 'acos' and 'acos' not in mapping:
            mapping['acos'] = col
        elif header == 'tacos' and 'tacos' not in mapping:
            mapping['tacos'] = col
        elif 'conversion' in header and 'conversion_rate' not in mapping:
            mapping['conversion_rate'] = col

    if 'sales' not in mapping and fallback_sales is not None:
        mapping['sales'] = fallback_sales
    return mapping


def _find_header_and_metrics(ws, block: dict[str, int | str]) -> tuple[int, dict[str, int]] | None:
    start = int(block['header_anchor_row']) + 1
    end = min(start + 6, ws.max_row)
    for row in range(start, end + 1):
        mm = _metric_map_for_row(ws, row, int(block['start_col']), int(block['end_col']))
        if 'spend' in mm and 'sales' in mm and ('impressions' in mm or 'clicks' in mm):
            return row, mm
    for row in range(start, end + 1):
        mm = _metric_map_for_row(ws, row, int(block['start_col']), int(block['end_col']))
        if 'spend' in mm and 'sales' in mm:
            return row, mm
    return None


def _is_subtotal_or_total(value: str) -> bool:
    normalized = _normalize_text(value)
    return 'subtotal' in normalized or normalized == 'total' or 'grand total' in normalized


def _ensure_weekly_columns(cursor) -> None:
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS weekly_ads (
            id SERIAL PRIMARY KEY,
            user_id INTEGER,
            sku VARCHAR(120),
            week_range VARCHAR(80),
            spend DOUBLE PRECISION,
            sales DOUBLE PRECISION
        )
        """
    )
    column_defs = [
        ('asin', 'VARCHAR(32)'),
        ('category', 'VARCHAR(200)'),
        ('impressions', 'DOUBLE PRECISION'),
        ('clicks', 'DOUBLE PRECISION'),
        ('page_views', 'DOUBLE PRECISION'),
        ('sessions', 'DOUBLE PRECISION'),
        ('ctr', 'DOUBLE PRECISION'),
        ('total_units', 'DOUBLE PRECISION'),
        ('total_sales', 'DOUBLE PRECISION'),
        ('acos', 'DOUBLE PRECISION'),
        ('tacos', 'DOUBLE PRECISION'),
        ('conversion_rate', 'DOUBLE PRECISION'),
    ]
    for col, col_type in column_defs:
        cursor.execute(f"ALTER TABLE weekly_ads ADD COLUMN IF NOT EXISTS {col} {col_type}")


def parse_weekly_ads_excel(file_path: Path, user_id: int | None = None) -> tuple[int, str]:
    wb = None
    try:
        wb = load_workbook(filename=file_path, data_only=True)
        ws = wb.active
        blocks = _detect_week_blocks(ws)
        if not blocks:
            return 0, 'Week header not found. Keep week labels like "15 Mar - 21 Mar".'

        aggregated: dict[str, dict[str, Any]] = {}
        weeks: set[str] = set()

        for block in blocks:
            header = _find_header_and_metrics(ws, block)
            if not header:
                continue
            header_row, metric_map = header
            week_range = str(block['week_range'])
            weeks.add(week_range)

            empty_streak = 0
            for row in range(header_row + 1, ws.max_row + 1):
                sku = str(_cell(ws, row, 1) or '').strip()
                asin = str(_cell(ws, row, 2) or '').strip()
                category = str(_cell(ws, row, 3) or '').strip()

                impressions = _to_number(_cell(ws, row, metric_map.get('impressions', 0)))
                clicks = _to_number(_cell(ws, row, metric_map.get('clicks', 0)))
                page_views = _to_number(_cell(ws, row, metric_map.get('page_views', 0)))
                sessions = _to_number(_cell(ws, row, metric_map.get('sessions', 0)))
                spend = _to_number(_cell(ws, row, metric_map.get('spend', 0)))
                sales = _to_number(_cell(ws, row, metric_map.get('sales', 0)))
                total_units = _to_number(_cell(ws, row, metric_map.get('total_units', 0)))
                total_sales = _to_number(_cell(ws, row, metric_map.get('total_sales', 0)))

                has_metrics = any(v > 0 for v in [impressions, clicks, page_views, sessions, spend, sales, total_units, total_sales])
                if not sku and not asin and not has_metrics:
                    empty_streak += 1
                    if empty_streak >= 40:
                        break
                    continue
                empty_streak = 0

                if (not sku and not asin) or _is_subtotal_or_total(sku):
                    continue

                key_label = sku or asin
                key = f"{week_range.lower()}|{key_label.lower()}|{asin.lower()}|{category.lower()}"
                if key not in aggregated:
                    aggregated[key] = {
                        'sku': sku or asin,
                        'asin': asin,
                        'category': category,
                        'week_range': week_range,
                        'impressions': 0.0,
                        'clicks': 0.0,
                        'page_views': 0.0,
                        'sessions': 0.0,
                        'spend': 0.0,
                        'sales': 0.0,
                        'total_units': 0.0,
                        'total_sales': 0.0,
                    }
                aggregated[key]['impressions'] += impressions
                aggregated[key]['clicks'] += clicks
                aggregated[key]['page_views'] += page_views
                aggregated[key]['sessions'] += sessions
                aggregated[key]['spend'] += spend
                aggregated[key]['sales'] += sales
                aggregated[key]['total_units'] += total_units
                aggregated[key]['total_sales'] += total_sales

        if not aggregated:
            return 0, 'No valid SKU rows found. Subtotal/total rows were skipped.'

        with transaction() as connection:
            with connection.cursor() as cursor:
                _ensure_weekly_columns(cursor)
                if user_id is not None:
                    cursor.execute("DELETE FROM weekly_ads WHERE user_id = %s", (user_id,))
                else:
                    cursor.execute('DELETE FROM weekly_ads')

                insert_sql = """
                    INSERT INTO weekly_ads (
                        user_id, sku, asin, category, week_range,
                        impressions, clicks, page_views, sessions, ctr,
                        spend, sales, total_units, total_sales, acos, tacos, conversion_rate
                    ) VALUES (
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s, %s
                    )
                """

                insert_values: list[tuple[Any, ...]] = []
                for row in aggregated.values():
                    impressions = float(row['impressions'])
                    clicks = float(row['clicks'])
                    sessions = float(row['sessions'])
                    spend = float(row['spend'])
                    sales = float(row['sales'])
                    total_sales = float(row['total_sales'])
                    total_units = float(row['total_units'])

                    ctr = (clicks / impressions * 100.0) if impressions > 0 else 0.0
                    acos = (spend / sales * 100.0) if sales > 0 else 0.0
                    tacos = (spend / total_sales * 100.0) if total_sales > 0 else 0.0
                    conversion_rate = (total_units / sessions * 100.0) if sessions > 0 else 0.0

                    insert_values.append(
                        (
                            user_id,
                            row['sku'],
                            row['asin'],
                            row['category'],
                            row['week_range'],
                            impressions,
                            clicks,
                            float(row['page_views']),
                            sessions,
                            round(ctr, 4),
                            spend,
                            sales,
                            total_units,
                            total_sales,
                            round(acos, 4),
                            round(tacos, 4),
                            round(conversion_rate, 4),
                        )
                    )

                cursor.executemany(insert_sql, insert_values)

        return len(aggregated), f'WoW upload successful. Weeks: {len(weeks)} | Rows: {len(aggregated)}.'
    except Exception as exc:
        return 0, str(exc)
    finally:
        if wb is not None:
            wb.close()
