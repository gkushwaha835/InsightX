from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Iterable

from flask import Response
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas


def _safe_filename(prefix: str, extension: str) -> str:
    stamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    return f'{prefix}_{stamp}.{extension}'


def export_dashboard_excel(
    page_title: str,
    overview: dict,
    trend: list[dict],
    analysis_rows: list[dict],
) -> Response:
    workbook = Workbook()
    ws_summary = workbook.active
    ws_summary.title = 'Summary'

    ws_summary.append([page_title])
    ws_summary.append([])
    ws_summary.append(['Total Spend', overview.get('total_spend', 0)])
    ws_summary.append(['Ads Sales', overview.get('ads_sales', 0)])
    ws_summary.append(['Total Sales', overview.get('total_sales', 0)])
    ws_summary.append(['ACOS', overview.get('acos', 0)])
    ws_summary.append(['TACOS', overview.get('tacos', 0)])
    ws_summary.append(['CTR', overview.get('ctr', 0)])
    ws_summary.append(['Conversion', overview.get('conversion_rate', 0)])

    ws_trend = workbook.create_sheet('Trend')
    ws_trend.append(['Month', 'Spend', 'Ads Sales', 'Total Sales', 'CTR', 'Efficiency'])
    for row in trend:
        ws_trend.append([
            row.get('month', ''),
            row.get('spend', 0),
            row.get('ads_sales', 0),
            row.get('total_sales', 0),
            row.get('ctr', 0),
            row.get('acos', 0),
        ])

    ws_analysis = workbook.create_sheet('Analysis')
    ws_analysis.append(['Entity', 'Spend', 'Ads Sales', 'Total Sales', 'Efficiency', 'CTR', 'Conversion', 'Status'])
    for row in analysis_rows:
        ws_analysis.append([
            row.get('label', ''),
            row.get('spend', 0),
            row.get('ads_sales', 0),
            row.get('total_sales', 0),
            row.get('efficiency', 0),
            row.get('ctr', 0),
            row.get('conversion_rate', 0),
            row.get('status', ''),
        ])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={_safe_filename("dashboard_export", "xlsx")}'},
    )


def _draw_table(pdf: canvas.Canvas, title: str, headers: Iterable[str], rows: list[list], y_start: int) -> int:
    y = y_start
    pdf.setFont('Helvetica-Bold', 11)
    pdf.drawString(30, y, title)
    y -= 16

    pdf.setFont('Helvetica-Bold', 8)
    x_positions = [30, 140, 220, 310, 400, 490, 560, 650]
    for idx, head in enumerate(headers):
        if idx < len(x_positions):
            pdf.drawString(x_positions[idx], y, str(head)[:20])
    y -= 12

    pdf.setFont('Helvetica', 8)
    for row in rows:
        for idx, value in enumerate(row):
            if idx < len(x_positions):
                pdf.drawString(x_positions[idx], y, str(value)[:22])
        y -= 11
        if y < 40:
            pdf.showPage()
            pdf.setFont('Helvetica', 8)
            y = 550
    return y


def export_dashboard_pdf(
    page_title: str,
    overview: dict,
    trend: list[dict],
    analysis_rows: list[dict],
) -> Response:
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=landscape(A4))

    pdf.setFont('Helvetica-Bold', 16)
    pdf.drawString(30, 570, page_title)
    pdf.setFont('Helvetica', 10)
    pdf.drawString(
        30,
        552,
        f"Spend: {overview.get('total_spend', 0)} | Ads Sales: {overview.get('ads_sales', 0)} | Total Sales: {overview.get('total_sales', 0)} | ACOS: {overview.get('acos', 0)}%",
    )

    y = 530
    trend_rows = [
        [r.get('month', ''), r.get('spend', 0), r.get('ads_sales', 0), r.get('total_sales', 0), r.get('ctr', 0), r.get('acos', 0)]
        for r in trend
    ]
    y = _draw_table(pdf, 'Trend', ['Month', 'Spend', 'Ads Sales', 'Total Sales', 'CTR', 'Efficiency'], trend_rows[:22], y)

    analysis_table_rows = [
        [
            r.get('label', ''),
            r.get('spend', 0),
            r.get('ads_sales', 0),
            r.get('total_sales', 0),
            r.get('efficiency', 0),
            r.get('ctr', 0),
            r.get('conversion_rate', 0),
            r.get('status', ''),
        ]
        for r in analysis_rows
    ]
    _draw_table(
        pdf,
        'Analysis Rows',
        ['Entity', 'Spend', 'Ads Sales', 'Total Sales', 'Efficiency', 'CTR', 'Conversion', 'Status'],
        analysis_table_rows[:40],
        y - 14,
    )

    pdf.save()
    buffer.seek(0)

    return Response(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename={_safe_filename("dashboard_export", "pdf")}'},
    )
