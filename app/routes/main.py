from __future__ import annotations

import csv
import json
import mimetypes
import re
import secrets
import smtplib
from datetime import datetime, timedelta
from email.message import EmailMessage
from io import StringIO
from pathlib import Path
from types import SimpleNamespace
from urllib.parse import urlencode
from urllib import error as urllib_error
from urllib import request as urllib_request
from uuid import uuid4

from flask import Blueprint, Response, abort, current_app, flash, redirect, render_template, request, send_file, session, url_for
from openpyxl import load_workbook

from ..auth import can_create_users, can_manage_users, can_view_user_data, get_current_user, login_required
from ..db import execute, fetch_all, fetch_one, get_app_setting, transaction
from ..services.report_history import load_history
from ..services.upload_service import _load_business_metrics, _parse_bulk_ads_metrics
from ..services.user_flow import log_user_login_activity, log_user_logout_activity

main_bp = Blueprint('main', __name__)


def get_setting(key: str, default: str = '') -> str:
    return get_app_setting(key, default)


def _row_to_namespace(row: dict | None) -> SimpleNamespace | None:
    if not row:
        return None
    return SimpleNamespace(**row)


def _get_log_or_404(log_id: int) -> SimpleNamespace:
    row = fetch_one("SELECT * FROM user_data_logs WHERE id = %s", (log_id,))
    if not row:
        abort(404)
    return SimpleNamespace(**row)


def latest_history_item(path: Path) -> dict | None:
    manifest = load_history(path)
    items = manifest.get('items', [])
    if not isinstance(items, list) or not items:
        return None
    for item in reversed(items):
        if isinstance(item, dict):
            return item
    return None


def latest_history_item_from_candidates(paths: list[Path]) -> dict | None:
    combined_items: list[dict] = []

    for path in paths:
        manifest = load_history(path)
        items = manifest.get('items', [])
        if not isinstance(items, list):
            continue
        for item in items:
            if isinstance(item, dict):
                combined_items.append(item)

    if not combined_items:
        return None

    return combined_items[-1]


def _as_bool(value: object, default: bool = False) -> bool:
    if value is None:
        return default
    return str(value).strip().lower() in {'1', 'true', 'yes', 'on'}


def _smtp_settings() -> dict[str, object]:
    raw_port = str(current_app.config.get('SMTP_PORT') or '587').strip()
    try:
        smtp_port = int(raw_port)
    except ValueError:
        smtp_port = 587
    if smtp_port <= 0:
        smtp_port = 587

    return {
        'host': str(current_app.config.get('SMTP_HOST') or '').strip(),
        'port': smtp_port,
        'username': str(current_app.config.get('SMTP_USERNAME') or '').strip(),
        'password': str(current_app.config.get('SMTP_PASSWORD') or '').strip(),
        'from_email': str(current_app.config.get('SMTP_FROM_EMAIL') or '').strip(),
        'from_name': str(current_app.config.get('SMTP_FROM_NAME') or 'InsightX').strip(),
        'use_tls': _as_bool(current_app.config.get('SMTP_USE_TLS'), default=True),
        'use_ssl': _as_bool(current_app.config.get('SMTP_USE_SSL'), default=False),
    }


def _smtp_configured(settings: dict[str, object]) -> bool:
    return bool(settings.get('host') and settings.get('from_email'))


def _google_oauth_settings() -> dict[str, str]:
    return {
        'client_id': str(current_app.config.get('GOOGLE_CLIENT_ID') or '').strip(),
        'client_secret': str(current_app.config.get('GOOGLE_CLIENT_SECRET') or '').strip(),
    }


def _google_oauth_configured() -> bool:
    settings = _google_oauth_settings()
    return bool(settings['client_id'] and settings['client_secret'])


def _exchange_google_code(code: str, redirect_uri: str) -> dict[str, object]:
    settings = _google_oauth_settings()
    payload = urlencode(
        {
            'code': code,
            'client_id': settings['client_id'],
            'client_secret': settings['client_secret'],
            'redirect_uri': redirect_uri,
            'grant_type': 'authorization_code',
        }
    ).encode('utf-8')
    request_obj = urllib_request.Request(
        'https://oauth2.googleapis.com/token',
        data=payload,
        headers={'Content-Type': 'application/x-www-form-urlencoded'},
        method='POST',
    )
    with urllib_request.urlopen(request_obj, timeout=15) as response:
        return json.loads(response.read().decode('utf-8'))


def _fetch_google_userinfo(access_token: str) -> dict[str, object]:
    request_obj = urllib_request.Request(
        'https://www.googleapis.com/oauth2/v2/userinfo',
        headers={'Authorization': f'Bearer {access_token}'},
    )
    with urllib_request.urlopen(request_obj, timeout=15) as response:
        return json.loads(response.read().decode('utf-8'))


def _generate_temporary_password(length: int = 12) -> str:
    safe_length = max(length, 8)
    alphabet = 'ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnopqrstuvwxyz23456789'
    return ''.join(secrets.choice(alphabet) for _ in range(safe_length))


def _load_brand_name_for_email(email: str) -> str:
    email = str(email or '').strip().lower()
    if not email:
        return ''
    row = fetch_one(
        """
        SELECT brand_name
        FROM user_identity
        WHERE lower(user_email) = %s
        LIMIT 1
        """,
        (email,),
    )
    if row:
        return str(row.get('brand_name') or '').strip()
    return ''


def _clean_display_name(value: str) -> str:
    raw = str(value or '').strip()
    if not raw:
        return ''
    if '@' in raw:
        raw = raw.split('@', 1)[0]
    for token in ('_', '.', '-', '+'):
        raw = raw.replace(token, ' ')
    parts = [part for part in raw.split() if part]
    cleaned_parts: list[str] = []
    for part in parts:
        match = re.match(r'^[A-Za-z]+', part)
        if match:
            cleaned_parts.append(match.group(0))
            break
    if not cleaned_parts and parts:
        cleaned_parts.append(parts[0])
    cleaned = ' '.join(cleaned_parts)
    if not cleaned:
        return ''
    return cleaned.title()


def _resolve_display_name(user: SimpleNamespace | None, fallback_email: str = '') -> str:
    if not user:
        return _clean_display_name(fallback_email) or 'User'
    role = str(getattr(user, 'role', '') or '').strip().lower()
    if role == 'admin':
        return 'Admin'
    username = str(getattr(user, 'username', '') or '').strip()
    display_name = _clean_display_name(username)
    if display_name:
        return display_name
    display_name = _clean_display_name(fallback_email)
    if display_name:
        return display_name
    return 'User'


def _read_email_logo() -> tuple[bytes | None, str]:
    logo_path = Path(__file__).resolve().parents[1] / 'static' / 'images' / 'insightx-logo.png'
    if not logo_path.exists():
        return None, 'png'
    try:
        logo_bytes = logo_path.read_bytes()
    except OSError:
        return None, 'png'
    mime_type, _ = mimetypes.guess_type(str(logo_path))
    if mime_type and '/' in mime_type:
        return logo_bytes, mime_type.split('/', 1)[1].lower()
    return logo_bytes, 'png'


def _send_temporary_password_email(recipient_email: str, temporary_password: str) -> None:
    settings = _smtp_settings()
    if not _smtp_configured(settings):
        raise RuntimeError('SMTP settings are not configured.')

    host = str(settings['host'])
    port = int(settings['port'])
    username = str(settings.get('username') or '')
    password = str(settings.get('password') or '')
    from_email = str(settings['from_email'])
    from_name = str(settings.get('from_name') or 'InsightX')
    use_tls = bool(settings.get('use_tls'))
    use_ssl = bool(settings.get('use_ssl'))
    logo_bytes, logo_subtype = _read_email_logo()
    logo_cid = 'insightx-logo'

    message = EmailMessage()
    message['Subject'] = 'InsightX Temporary Password'
    message['From'] = f'{from_name} <{from_email}>'
    message['To'] = recipient_email
    message.set_content(
        '\n'.join(
            [
                'Hello,',
                '',
                'A password reset request was received for your InsightX account.',
                f'Temporary password: {temporary_password}',
                '',
                'Please login with this password and immediately change it from Settings.',
                '',
                'If you did not request this, please contact support.',
            ]
        )
    )
    logo_block = ''
    if logo_bytes:
        logo_block = (
            '<table role="presentation" cellpadding="0" cellspacing="0" border="0" '
            'style="margin-top:20px;">'
            '<tr>'
            '<td style="padding-right:12px;vertical-align:middle;">'
            f'<img src="cid:{logo_cid}" alt="InsightX" '
            'style="display:block;width:165px;max-width:165px;height:auto;border:0;outline:none;text-decoration:none;" />'
            '</td>'
            '<td style="vertical-align:middle;color:#64748b;font-size:13px;line-height:1.4;">'
            '<div style="font-weight:700;color:#0f172a;margin-bottom:2px;">InsightX</div>'
            '<div>Data • Insight • Impact</div>'
            '</td>'
            '</tr>'
            '</table>'
        )
    message.add_alternative(
        (
            '<html><body style="font-family:Segoe UI,Tahoma,Verdana,sans-serif;color:#0f172a;">'
            '<p>Hello,</p>'
            '<p>A password reset request was received for your InsightX account.</p>'
            f'<p><b>Temporary password:</b> {temporary_password}</p>'
            '<p>Please login with this password and immediately change it from Settings.</p>'
            '<p>If you did not request this, please contact support.</p>'
            f'{logo_block}'
            '</body></html>'
        ),
        subtype='html',
    )
    if logo_bytes:
        html_part = message.get_payload()[-1]
        html_part.add_related(
            logo_bytes,
            maintype='image',
            subtype=logo_subtype,
            cid=f'<{logo_cid}>',
            disposition='inline',
        )

    if use_ssl:
        with smtplib.SMTP_SSL(host, port, timeout=15) as smtp:
            if username:
                smtp.login(username, password)
            smtp.send_message(message)
        return

    with smtplib.SMTP(host, port, timeout=15) as smtp:
        smtp.ehlo()
        if use_tls:
            smtp.starttls()
            smtp.ehlo()
        if username:
            smtp.login(username, password)
        smtp.send_message(message)


def _reset_password_and_send_email(email: str) -> bool:
    account = fetch_one(
        """
        SELECT id, username, session_version
        FROM users
        WHERE lower(username) = %s
        LIMIT 1
        """,
        (str(email or '').strip().lower(),),
    )
    if not account:
        return False

    temporary_password = _generate_temporary_password()
    next_session_version = int(account.get('session_version') or 1) + 1

    with transaction() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                UPDATE users
                SET password = %s, session_version = %s, updated_at = NOW()
                WHERE id = %s
                """,
                (temporary_password, next_session_version, account['id']),
            )
            _send_temporary_password_email(str(account.get('username') or email), temporary_password)
    return True


def _set_authenticated_session(account: SimpleNamespace) -> None:
    session.clear()
    session['user'] = account.username
    session['login_email'] = account.username.lower()
    session['login_phone'] = ''
    session['account_username'] = account.username
    session['user_role'] = account.role
    session['can_create_users'] = bool(account.can_create_users)
    session['session_version'] = int(account.session_version)
    session['is_guest_user'] = False
    session['show_brand_name_modal'] = False
    session['member_visit_id'] = uuid4().hex


@main_bp.route('/auth/google')
def google_login():
    if session.get('user'):
        return redirect(url_for('main.home'))

    auth_mode = (request.args.get('mode') or 'login').strip().lower()
    if auth_mode not in {'login', 'signup'}:
        auth_mode = 'login'

    settings = _google_oauth_settings()
    if not _google_oauth_configured():
        flash('Google login is not configured yet. Please try again later.', 'danger')
        return redirect(url_for('main.login'))

    state = secrets.token_urlsafe(24)
    session['google_oauth_state'] = state
    session['google_oauth_mode'] = auth_mode
    redirect_uri = url_for('main.google_callback', _external=True)
    query = urlencode(
        {
            'client_id': settings['client_id'],
            'redirect_uri': redirect_uri,
            'response_type': 'code',
            'scope': 'openid email profile',
            'state': state,
            'access_type': 'online',
            'prompt': 'select_account',
        }
    )
    return redirect(f'https://accounts.google.com/o/oauth2/auth?{query}')


@main_bp.route('/auth/google/callback')
def google_callback():
    if session.get('user'):
        return redirect(url_for('main.home'))

    if not _google_oauth_configured():
        flash('Google login is not configured yet. Please try again later.', 'danger')
        return redirect(url_for('main.login'))

    error_message = request.args.get('error')
    if error_message:
        flash('Google login was cancelled or denied.', 'warning')
        return redirect(url_for('main.login'))

    expected_state = str(session.pop('google_oauth_state', '') or '')
    auth_mode = str(session.pop('google_oauth_mode', 'login') or 'login').strip().lower()
    if auth_mode not in {'login', 'signup'}:
        auth_mode = 'login'
    received_state = str(request.args.get('state') or '')
    if not expected_state or expected_state != received_state:
        flash('Google login could not be verified. Please try again.', 'danger')
        return redirect(url_for('main.login'))

    code = str(request.args.get('code') or '')
    if not code:
        flash('Google login did not return an authorization code.', 'danger')
        return redirect(url_for('main.login'))

    try:
        token_data = _exchange_google_code(code, url_for('main.google_callback', _external=True))
        access_token = str(token_data.get('access_token') or '')
        if not access_token:
            raise RuntimeError('Google token exchange did not return an access token.')
        profile = _fetch_google_userinfo(access_token)
    except Exception as exc:  # noqa: BLE001
        current_app.logger.exception('Google login failed: %s', exc)
        flash('Google login failed. Please try again.', 'danger')
        return redirect(url_for('main.login'))

    email = str(profile.get('email') or '').strip().lower()
    if not email:
        flash('Google account did not return an email address.', 'danger')
        return redirect(url_for('main.login'))
    if not bool(profile.get('verified_email') or profile.get('email_verified')):
        flash('Google account email is not verified.', 'danger')
        return redirect(url_for('main.login'))

    account = _row_to_namespace(
        fetch_one(
            """
            SELECT id, username, password, role, can_create_users, session_version
            FROM users
            WHERE lower(username) = %s
            LIMIT 1
            """,
            (email,),
        )
    )

    if not account:
        execute(
            """
            INSERT INTO users (username, password, role, can_create_users, session_version, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, NOW(), NOW())
            """,
            (email, uuid4().hex, 'user', False, 1),
        )
        account = _row_to_namespace(
            fetch_one(
                """
                SELECT id, username, password, role, can_create_users, session_version
                FROM users
                WHERE lower(username) = %s
                LIMIT 1
                """,
                (email,),
            )
        )

    if not account:
        flash('Could not create your account right now. Please try again.', 'danger')
        return redirect(url_for('main.login'))

    _set_authenticated_session(account)
    log_user_login_activity()
    if auth_mode == 'signup':
        flash('Signed up with Google successfully.', 'success')
    else:
        flash('Logged in with Google successfully.', 'success')
    return redirect(url_for('main.home'))


@main_bp.route('/', methods=['GET'])
@main_bp.route('/index.php', methods=['GET'])
@main_bp.route('/info', methods=['GET'])
@main_bp.route('/info.html', methods=['GET'])
def info():
    if session.get('user'):
        return redirect(url_for('main.home'))
    return render_template('info.html')


@main_bp.route('/blog/lower-acos', methods=['GET'])
@main_bp.route('/blog/lower-acos.html', methods=['GET'])
def blog_lower_acos():
    return render_template('blog-lower-acos.html')


@main_bp.route('/blog/tacos-vs-acos', methods=['GET'])
@main_bp.route('/blog/tacos-vs-acos.html', methods=['GET'])
def blog_tacos_vs_acos():
    return render_template('blog-tacos-vs-acos.html')


@main_bp.route('/blog/amazon-updates', methods=['GET'])
@main_bp.route('/blog/amazon-updates.html', methods=['GET'])
def blog_amazon_updates():
    return render_template('blog-amazon-updates.html')


@main_bp.route('/login', methods=['GET', 'POST'])
@main_bp.route('/login.php', methods=['GET', 'POST'])
@main_bp.route('/login.html', methods=['GET', 'POST'])
def login():
    selected_auth_mode = 'login'

    if request.method == 'POST':
        email_raw = (request.form.get('email') or '').strip()
        email = email_raw.lower()
        password = (request.form.get('password') or '').strip()
        selected_auth_mode = (request.form.get('auth_mode') or 'login').strip().lower()

        if selected_auth_mode == 'signup':
            confirm_password = (request.form.get('confirm_password') or '').strip()
            if not email or not password:
                flash('Email and password are required.', 'danger')
            elif '@' not in email:
                flash('Please enter a valid email address.', 'danger')
            elif len(password) < 6:
                flash('Password must be at least 6 characters.', 'danger')
            elif password != confirm_password:
                flash('Password and confirm password do not match.', 'danger')
            elif fetch_one("SELECT id FROM users WHERE lower(username) = %s LIMIT 1", (email,)):
                flash('Account already exists. Please login.', 'warning')
            else:
                execute(
                    """
                    INSERT INTO users (username, password, role, can_create_users, session_version, created_at, updated_at)
                    VALUES (%s, %s, %s, %s, %s, NOW(), NOW())
                    """,
                    (email, password, 'user', False, 1),
                )
                account = _row_to_namespace(
                    fetch_one(
                        """
                        SELECT id, username, password, role, can_create_users, session_version
                        FROM users
                        WHERE lower(username) = %s
                        LIMIT 1
                        """,
                        (email,),
                    )
                )
                if account:
                    _set_authenticated_session(account)
                    log_user_login_activity()
                    flash('Signup successful. Welcome!', 'success')
                    return redirect(url_for('main.home'))
                flash('Signup completed, but auto-login failed. Please login.', 'warning')
                selected_auth_mode = 'login'
        else:
            selected_auth_mode = 'login'
            if not email or not password:
                flash('Email and password are required.', 'danger')
            else:
                account = _row_to_namespace(
                    fetch_one(
                        """
                        SELECT id, username, password, role, can_create_users, session_version
                        FROM users
                        WHERE lower(username) = %s
                        LIMIT 1
                        """,
                        (email,),
                    )
                )
                if account and account.password == password:
                    _set_authenticated_session(account)
                    log_user_login_activity()
                    return redirect(url_for('main.home'))
                flash('Invalid email or password.', 'danger')

    if session.get('user'):
        return redirect(url_for('main.home'))

    return render_template('login.html', auth_mode=selected_auth_mode)


@main_bp.route('/forgot-password', methods=['GET', 'POST'])
@main_bp.route('/forgot-password.php', methods=['GET', 'POST'])
def forgot_password():
    if session.get('user'):
        return redirect(url_for('main.home'))

    if request.method == 'POST':
        email = (request.form.get('email') or '').strip().lower()
        if not email or '@' not in email:
            flash('Please enter a valid email address.', 'danger')
            return render_template('forgot_password.html')

        settings = _smtp_settings()
        if not _smtp_configured(settings):
            current_app.logger.warning('Forgot password requested but SMTP is not configured.')
            flash('Forgot password is not available right now. Please contact support.', 'danger')
            return render_template('forgot_password.html')

        try:
            _reset_password_and_send_email(email)
        except Exception as exc:  # noqa: BLE001
            current_app.logger.exception('Failed forgot-password flow for %s: %s', email, exc)
            flash('Could not send reset email right now. Please try again later.', 'danger')
            return render_template('forgot_password.html')

        flash(
            'If this email is registered, a temporary password has been sent. Please check your inbox.',
            'success',
        )
        return redirect(url_for('main.login'))

    return render_template('forgot_password.html')


@main_bp.route('/favicon.ico')
def favicon():
    return ('', 204)


@main_bp.route('/logout')
def logout():
    log_user_logout_activity()
    session.clear()
    return redirect(url_for('main.login'))


@main_bp.route('/home')
@login_required
def home():
    user = get_current_user()
    if user is None:
        return redirect(url_for('main.login'))
    brand_name = str(session.get('brand_name') or '').strip()
    if not brand_name:
        brand_name = _load_brand_name_for_email(str(session.get('login_email') or user.username))
        if brand_name:
            session['brand_name'] = brand_name
    display_name = _resolve_display_name(user, str(session.get('login_email') or user.username))
    return render_template(
        'home.html',
        user=user,
        display_name=display_name,
        show_brand_name_modal=False,
        brand_name=brand_name,
    )


@main_bp.route('/set-brand-name', methods=['POST'])
@login_required
def set_brand_name():
    brand_name = (request.form.get('brand_name') or '').strip()
    if not brand_name:
        flash('Please enter your brand name.', 'danger')
        session['show_brand_name_modal'] = True
        return redirect(url_for('main.home'))

    session['brand_name'] = brand_name
    email = str(session.get('login_email') or '').strip().lower()
    member_visit_id = str(session.get('member_visit_id') or '').strip()
    if email and member_visit_id:
        try:
            execute(
                """
                UPDATE user_data_logs
                SET brand_name = %s, updated_at = NOW()
                WHERE lower(user_email) = %s AND COALESCE(member_visit_id, '') = %s
                """,
                (brand_name, email, member_visit_id),
            )
            execute(
                """
                UPDATE user_identity
                SET brand_name = %s, updated_at = NOW()
                WHERE lower(user_email) = %s
                """,
                (brand_name, email),
            )
        except Exception as exc:  # noqa: BLE001
            current_app.logger.warning('Brand name saved in session only for %s: %s', email, exc)
            flash('Brand name saved for this session, but the database is unavailable right now.', 'warning')
    session['show_brand_name_modal'] = False
    return redirect(url_for('main.home'))


@main_bp.route('/skip-brand-name', methods=['POST'])
@login_required
def skip_brand_name():
    role = str(session.get('user_role') or 'user').strip().lower()
    if role in {'admin', 'co_admin', 'cd_admin'}:
        session['show_brand_name_modal'] = False
        return redirect(url_for('main.home'))
    flash('Brand name is required to continue.', 'danger')
    session['show_brand_name_modal'] = True
    return redirect(url_for('main.home'))


@main_bp.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    user = get_current_user()
    if user is None:
        return redirect(url_for('main.login'))

    errors: list[str] = []
    success = ''

    if request.method == 'POST':
        action = (request.form.get('action') or '').strip()

        if action == 'change_password' and not session.get('is_guest_user', False):
            current_password = (request.form.get('current_password') or '').strip()
            new_password = (request.form.get('new_password') or '').strip()
            confirm_password = (request.form.get('confirm_password') or '').strip()

            if user.password != current_password:
                errors.append('Current password is incorrect.')
            elif not new_password:
                errors.append('New password is required.')
            elif new_password != confirm_password:
                errors.append('New password and confirm password do not match.')
            else:
                next_version = int(user.session_version or 1) + 1
                execute(
                    """
                    UPDATE users
                    SET password = %s, session_version = %s, updated_at = NOW()
                    WHERE id = %s
                    """,
                    (new_password, next_version, user.id),
                )
                user.password = new_password
                user.session_version = next_version
                session['session_version'] = next_version
                success = 'Password updated successfully.'

        elif action == 'logout_other_devices' and not session.get('is_guest_user', False):
            next_version = int(user.session_version or 1) + 1
            execute(
                "UPDATE users SET session_version = %s, updated_at = NOW() WHERE id = %s",
                (next_version, user.id),
            )
            user.session_version = next_version
            session['session_version'] = next_version
            success = 'All other devices were logged out for this account.'

        elif action == 'add_user' and can_create_users(user):
            username = (request.form.get('username') or '').strip()
            password = (request.form.get('password') or '').strip()
            role = (request.form.get('role') or 'user').strip()
            allowed_roles = ['user', 'co_admin'] if can_manage_users(user) else ['user']

            if not username or not password:
                errors.append('Username and password are required.')
            elif role not in allowed_roles:
                errors.append('Selected role is not allowed.')
            elif fetch_one("SELECT id FROM users WHERE username = %s LIMIT 1", (username,)):
                errors.append('That username already exists.')
            else:
                execute(
                    """
                    INSERT INTO users (username, password, role, can_create_users, session_version, created_at, updated_at)
                    VALUES (%s, %s, %s, %s, %s, NOW(), NOW())
                    """,
                    (username, password, role, False, 1),
                )
                success = 'User added successfully.' if role == 'user' else 'CO-Admin added successfully.'

        elif action == 'toggle_create_users' and can_manage_users(user):
            target_username = (request.form.get('target_username') or '').strip()
            grant_access = (request.form.get('grant_access') or '0').strip() == '1'
            target = _row_to_namespace(
                fetch_one(
                    """
                    SELECT id, username, role, can_create_users
                    FROM users
                    WHERE username = %s
                    LIMIT 1
                    """,
                    (target_username,),
                )
            )
            if not target:
                errors.append('Selected user was not found.')
            elif target.role not in ('co_admin', 'cd_admin'):
                errors.append('Only CO-Admins can receive this permission.')
            else:
                execute(
                    "UPDATE users SET can_create_users = %s, updated_at = NOW() WHERE id = %s",
                    (grant_access, target.id),
                )
                success = 'CO-Admin permission updated.'

        elif action == 'delete_user' and can_manage_users(user):
            target_username = (request.form.get('target_username') or '').strip()
            target = _row_to_namespace(
                fetch_one(
                    """
                    SELECT id, username, role
                    FROM users
                    WHERE username = %s
                    LIMIT 1
                    """,
                    (target_username,),
                )
            )
            if not target:
                errors.append('Selected user was not found.')
            elif target.username == user.username:
                errors.append('You cannot delete your own account.')
            elif target.role == 'admin':
                errors.append('Admin accounts cannot be deleted.')
            else:
                execute("DELETE FROM users WHERE id = %s", (target.id,))
                success = 'User removed successfully.'

    users = []
    if can_create_users(user) or can_manage_users(user):
        users = [
            SimpleNamespace(**row)
            for row in fetch_all(
                """
                SELECT id, username, role, can_create_users, session_version
                FROM users
                ORDER BY username ASC
                """
            )
        ]

    return render_template(
        'settings.html',
        user=user,
        users=users,
        can_manage=can_manage_users(user),
        can_create=can_create_users(user),
        can_view_data=can_view_user_data(user),
        brand_name=str(session.get('brand_name') or ''),
        is_guest_user=bool(session.get('is_guest_user', False)),
        errors=errors,
        success=success,
    )


def _safe_log_file_path(log: SimpleNamespace) -> Path:
    candidate = Path(log.file_path).resolve()
    app_root = Path(__file__).resolve().parents[2]
    upload_roots = [
        (app_root / 'uploads').resolve(),
        app_root.resolve(),
    ]
    if not any(str(candidate).startswith(str(root)) for root in upload_roots):
        raise FileNotFoundError('File path is not allowed.')
    if not candidate.exists():
        raise FileNotFoundError('File not found.')
    return candidate


def _is_allowed_upload_path(candidate: Path) -> bool:
    app_root = Path(__file__).resolve().parents[2]
    upload_roots = [
        (app_root / 'uploads').resolve(),
        app_root.resolve(),
    ]
    return any(str(candidate).startswith(str(root)) for root in upload_roots)


def _preview_file_rows(file_path: Path, max_rows: int = 50, max_cols: int = 20) -> tuple[list[str], list[list[str]]]:
    suffix = file_path.suffix.lower()
    if suffix == '.csv':
        with file_path.open('r', encoding='utf-8-sig', newline='') as handle:
            reader = csv.reader(handle)
            rows = [row[:max_cols] for row in reader]
    else:
        workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
        try:
            ws = workbook.active
            rows = []
            for row in ws.iter_rows(min_row=1, max_row=max_rows + 1, values_only=True):
                rows.append([str(cell) if cell is not None else '' for cell in list(row)[:max_cols]])
        finally:
            workbook.close()

    if not rows:
        return [], []

    header = rows[0]
    body = rows[1:max_rows + 1]
    return header, body


def _format_duration(start_at: datetime | None, end_at: datetime | None) -> str:
    if not start_at or not end_at:
        return '0m'
    total_seconds = int((end_at - start_at).total_seconds())
    if total_seconds < 0:
        total_seconds = 0

    days, rem = divmod(total_seconds, 86400)
    hours, rem = divmod(rem, 3600)
    minutes, _ = divmod(rem, 60)

    if days > 0:
        return f'{days}d {hours}h {minutes}m'
    if hours > 0:
        return f'{hours}h {minutes}m'
    if minutes > 0:
        return f'{minutes}m'
    return f'{rem}s'


def _normalize_option_key(value: str) -> str:
    return ' '.join(str(value or '').strip().lower().split())


def _is_management_account_email(email: str) -> bool:
    row = fetch_one(
        """
        SELECT role
        FROM users
        WHERE lower(username) = %s
        LIMIT 1
        """,
        (str(email or '').strip().lower(),),
    )
    if not row:
        return False
    role = str(row.get('role') or '').strip().lower()
    return role in {'admin', 'co_admin', 'cd_admin'}


def _format_duration_full_seconds(total_seconds: int) -> str:
    safe_seconds = max(int(total_seconds or 0), 0)
    days, rem = divmod(safe_seconds, 86400)
    hours, rem = divmod(rem, 3600)
    minutes, seconds = divmod(rem, 60)

    if days > 0:
        return f'{days}d {hours}h {minutes}m {seconds}s'
    if hours > 0:
        return f'{hours}h {minutes}m {seconds}s'
    if minutes > 0:
        return f'{minutes}m {seconds}s'
    return f'{seconds}s'


def _parse_users_data_filters(args) -> dict:
    preset = str(args.get('preset') or 'all').strip().lower()
    if preset not in {'all', 'today', '7d', '30d', 'custom'}:
        preset = 'all'

    date_from_input = str(args.get('date_from') or '').strip()
    date_to_input = str(args.get('date_to') or '').strip()

    now = datetime.now()
    today_start = datetime.combine(now.date(), datetime.min.time())
    start_at: datetime | None = None
    end_at: datetime | None = None
    date_from_value = ''
    date_to_value = ''

    if preset == 'today':
        start_at = today_start
        end_at = start_at + timedelta(days=1)
        date_from_value = start_at.date().isoformat()
        date_to_value = start_at.date().isoformat()
    elif preset == '7d':
        start_at = today_start - timedelta(days=6)
        end_at = today_start + timedelta(days=1)
        date_from_value = start_at.date().isoformat()
        date_to_value = (end_at - timedelta(days=1)).date().isoformat()
    elif preset == '30d':
        start_at = today_start - timedelta(days=29)
        end_at = today_start + timedelta(days=1)
        date_from_value = start_at.date().isoformat()
        date_to_value = (end_at - timedelta(days=1)).date().isoformat()
    elif preset == 'custom':
        if date_from_input:
            try:
                parsed_from = datetime.strptime(date_from_input, '%Y-%m-%d').date()
                start_at = datetime.combine(parsed_from, datetime.min.time())
                date_from_value = parsed_from.isoformat()
            except ValueError:
                start_at = None
                date_from_value = ''
        if date_to_input:
            try:
                parsed_to = datetime.strptime(date_to_input, '%Y-%m-%d').date()
                end_at = datetime.combine(parsed_to, datetime.min.time()) + timedelta(days=1)
                date_to_value = parsed_to.isoformat()
            except ValueError:
                end_at = None
                date_to_value = ''
        if start_at and end_at and end_at <= start_at:
            end_at = start_at + timedelta(days=1)
            date_to_value = (end_at - timedelta(days=1)).date().isoformat()

    search = str(args.get('q') or '').strip()
    sort_by = str(args.get('sort_by') or 'last_activity_at').strip().lower()
    sort_order = str(args.get('sort_order') or 'desc').strip().lower()
    if sort_order not in {'asc', 'desc'}:
        sort_order = 'desc'
    if sort_by not in {'user_email', 'brand_name', 'activity_count', 'first_activity_at', 'last_activity_at', 'usage_seconds', 'top_option_count'}:
        sort_by = 'last_activity_at'

    try:
        page = int(args.get('page', 1))
    except (TypeError, ValueError):
        page = 1
    if page < 1:
        page = 1

    try:
        per_page = int(args.get('per_page', 10))
    except (TypeError, ValueError):
        per_page = 10
    if per_page not in {10, 25, 50, 100}:
        per_page = 10

    return {
        'preset': preset,
        'date_from': date_from_value,
        'date_to': date_to_value,
        'start_at': start_at,
        'end_at': end_at,
        'q': search,
        'sort_by': sort_by,
        'sort_order': sort_order,
        'page': page,
        'per_page': per_page,
    }


def _fetch_users_data_rows(start_at: datetime | None, end_at: datetime | None) -> list[dict]:
    date_filter_sql = ''
    date_filter_params: list[datetime] = []
    if start_at is not None:
        date_filter_sql += " AND logs.created_at >= %s"
        date_filter_params.append(start_at)
    if end_at is not None:
        date_filter_sql += " AND logs.created_at < %s"
        date_filter_params.append(end_at)

    params: list = []
    if date_filter_params:
        params.extend(date_filter_params)  # display_logs
        params.extend(date_filter_params)  # session_bounds
        params.extend(date_filter_params)  # option_rank

    query = f"""
        WITH display_logs AS (
            SELECT
                logs.*,
                COALESCE(logs.member_visit_id, ('legacy-' || logs.id::text)) AS visit_key,
                lower(logs.user_email) AS email_key
            FROM user_data_logs logs
            WHERE NOT EXISTS (
                SELECT 1
                FROM users u
                WHERE lower(u.username) = lower(logs.user_email)
                  AND lower(u.role) IN ('admin', 'co_admin', 'cd_admin')
            )
              AND COALESCE(logs.option_used, '') NOT IN ('Login', 'Logout')
              {date_filter_sql}
        ),
        session_bounds AS (
            SELECT
                COALESCE(logs.member_visit_id, ('legacy-' || logs.id::text)) AS visit_key,
                lower(logs.user_email) AS email_key,
                MIN(logs.created_at) AS first_seen_at,
                MAX(logs.created_at) AS last_seen_at
            FROM user_data_logs logs
            WHERE NOT EXISTS (
                SELECT 1
                FROM users u
                WHERE lower(u.username) = lower(logs.user_email)
                  AND lower(u.role) IN ('admin', 'co_admin', 'cd_admin')
            )
              AND COALESCE(logs.option_used, '') NOT IN ('Login', 'Logout')
              {date_filter_sql}
            GROUP BY
                COALESCE(logs.member_visit_id, ('legacy-' || logs.id::text)),
                lower(logs.user_email)
        ),
        option_rank AS (
            SELECT
                COALESCE(logs.member_visit_id, ('legacy-' || logs.id::text)) AS visit_key,
                lower(logs.user_email) AS email_key,
                COALESCE(NULLIF(logs.option_used, ''), 'Unknown') AS option_used,
                COUNT(*) AS option_count,
                ROW_NUMBER() OVER (
                    PARTITION BY COALESCE(logs.member_visit_id, ('legacy-' || logs.id::text)), lower(logs.user_email)
                    ORDER BY COUNT(*) DESC, MAX(logs.created_at) DESC
                ) AS rn
            FROM user_data_logs logs
            WHERE NOT EXISTS (
                SELECT 1
                FROM users u
                WHERE lower(u.username) = lower(logs.user_email)
                  AND lower(u.role) IN ('admin', 'co_admin', 'cd_admin')
            )
              AND COALESCE(logs.option_used, '') NOT IN ('Login', 'Logout')
              {date_filter_sql}
            GROUP BY
                COALESCE(logs.member_visit_id, ('legacy-' || logs.id::text)),
                lower(logs.user_email),
                COALESCE(NULLIF(logs.option_used, ''), 'Unknown')
        ),
        top_option AS (
            SELECT visit_key, email_key, option_used, option_count
            FROM option_rank
            WHERE rn = 1
        )
        SELECT
            COALESCE(ui.id, MIN(d.user_identity_id)) AS user_code,
            d.visit_key AS member_visit_id,
            d.email_key AS user_email,
            COALESCE(
                ui.latest_phone,
                (ARRAY_AGG(d.phone_number ORDER BY d.created_at DESC, d.id DESC))[1],
                ''
            ) AS phone_number,
            COALESCE(
                NULLIF((ARRAY_AGG(COALESCE(d.brand_name, '') ORDER BY d.created_at DESC, d.id DESC))[1], ''),
                ''
            ) AS brand_name,
            COUNT(d.id) AS activity_count,
            sb.first_seen_at AS first_activity_at,
            sb.last_seen_at AS last_activity_at,
            COALESCE(top.option_used, 'Unknown') AS top_option_used,
            COALESCE(top.option_count, 0) AS top_option_count
        FROM display_logs d
        LEFT JOIN user_identity ui
            ON lower(ui.user_email) = d.email_key
        LEFT JOIN session_bounds sb
            ON sb.visit_key = d.visit_key AND sb.email_key = d.email_key
        LEFT JOIN top_option top
            ON top.visit_key = d.visit_key AND top.email_key = d.email_key
        GROUP BY
            d.visit_key,
            d.email_key,
            ui.id,
            ui.latest_phone,
            ui.brand_name,
            sb.first_seen_at,
            sb.last_seen_at,
            top.option_used,
            top.option_count
        ORDER BY sb.last_seen_at DESC NULLS LAST
    """
    return fetch_all(query, tuple(params))


def _compute_user_status(last_activity_at: datetime | None) -> tuple[str, str]:
    if not last_activity_at:
        return 'offline', 'secondary'
    age_seconds = int((datetime.now() - last_activity_at).total_seconds())
    if age_seconds < 0:
        age_seconds = 0
    if age_seconds <= 300:
        return 'active', 'success'
    if age_seconds <= 3600:
        return 'idle', 'warning'
    return 'offline', 'secondary'


def _load_users_summary(args) -> tuple[list[dict], dict]:
    filters = _parse_users_data_filters(args)
    grouped = _fetch_users_data_rows(filters.get('start_at'), filters.get('end_at'))

    rows: list[dict] = []
    for row in grouped:
        first_activity_at = row.get('first_activity_at')
        last_activity_at = row.get('last_activity_at')
        usage_seconds = 0
        if first_activity_at and last_activity_at:
            usage_seconds = max(int((last_activity_at - first_activity_at).total_seconds()), 0)
        status_key, status_class = _compute_user_status(last_activity_at)
        rows.append(
            {
                'user_email': str(row.get('user_email') or ''),
                'member_visit_id': str(row.get('member_visit_id') or ''),
                'phone_number': str(row.get('phone_number') or ''),
                'brand_name': str(row.get('brand_name') or ''),
                'user_code': int(row.get('user_code') or 0),
                'activity_count': int(row.get('activity_count') or 0),
                'first_activity_at': first_activity_at,
                'last_activity_at': last_activity_at,
                'usage_seconds': usage_seconds,
                'usage_duration': _format_duration(first_activity_at, last_activity_at),
                'usage_duration_full': _format_duration_full_seconds(usage_seconds),
                'top_option_used': str(row.get('top_option_used') or 'Unknown'),
                'top_option_count': int(row.get('top_option_count') or 0),
                'status_key': status_key,
                'status_class': status_class,
            }
        )

    search = str(filters.get('q') or '').strip().lower()
    if search:
        rows = [
            row
            for row in rows
            if search in row['user_email'].lower()
            or search in row['phone_number'].lower()
            or search in row['brand_name'].lower()
            or search in row['top_option_used'].lower()
        ]

    sort_by = str(filters.get('sort_by') or 'last_activity_at')
    sort_order = str(filters.get('sort_order') or 'desc')
    reverse = sort_order == 'desc'

    def _sort_key(item: dict):
        if sort_by in {'user_email', 'brand_name'}:
            return str(item.get(sort_by) or '').lower()
        if sort_by in {'activity_count', 'usage_seconds', 'top_option_count'}:
            return int(item.get(sort_by) or 0)
        if sort_by in {'first_activity_at', 'last_activity_at'}:
            return item.get(sort_by) or datetime.min
        return item.get('last_activity_at') or datetime.min

    rows = sorted(rows, key=_sort_key, reverse=reverse)
    return rows, filters


def _ensure_soft_delete_table(cursor) -> None:
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS user_data_logs_deleted (
            id BIGSERIAL PRIMARY KEY,
            delete_batch_id TEXT NOT NULL,
            original_id BIGINT,
            deleted_at TIMESTAMP NOT NULL DEFAULT NOW(),
            deleted_by TEXT,
            delete_scope TEXT,
            log_data JSONB NOT NULL
        )
        """
    )


def _get_soft_delete_batch_count(batch_id: str) -> int:
    if not batch_id:
        return 0
    table_row = fetch_one("SELECT to_regclass('public.user_data_logs_deleted') AS reg") or {}
    if not table_row.get('reg'):
        return 0
    row = fetch_one(
        "SELECT COUNT(*) AS c FROM user_data_logs_deleted WHERE delete_batch_id = %s",
        (batch_id,),
    ) or {}
    return int(row.get('c') or 0)


def _build_ads_report_rows(bulk_path: Path, business_path: Path) -> tuple[list[dict], dict]:
    ads_metrics, _, _ = _parse_bulk_ads_metrics(bulk_path)
    business_metrics = _load_business_metrics(business_path)

    all_asins = sorted(set(ads_metrics.keys()) | set(business_metrics.keys()))
    rows: list[dict] = []

    for asin in all_asins:
        ads_row = ads_metrics.get(asin, {})
        business_row = business_metrics.get(asin, {})

        rows.append(
            {
                'asin': asin,
                'impressions': float(ads_row.get('impressions', 0.0)),
                'clicks': float(ads_row.get('clicks', 0.0)),
                'page_views': float(business_row.get('page_views', 0.0)),
                'sessions': float(business_row.get('sessions', 0.0)),
                'spend': float(ads_row.get('spend', 0.0)),
                'sales': float(ads_row.get('sales', 0.0)),
                'units_ordered': float(business_row.get('total_units', 0.0)),
                'ordered_product_sales': float(business_row.get('total_sales', 0.0)),
            }
        )

    rows.sort(key=lambda r: (r['ordered_product_sales'], r['sales'], r['spend']), reverse=True)

    totals = {
        'impressions': sum(r['impressions'] for r in rows),
        'clicks': sum(r['clicks'] for r in rows),
        'page_views': sum(r['page_views'] for r in rows),
        'sessions': sum(r['sessions'] for r in rows),
        'spend': sum(r['spend'] for r in rows),
        'sales': sum(r['sales'] for r in rows),
        'units_ordered': sum(r['units_ordered'] for r in rows),
        'ordered_product_sales': sum(r['ordered_product_sales'] for r in rows),
    }
    return rows, totals


@main_bp.route('/settings/users-data')
@login_required
def users_data():
    user = get_current_user()
    if not can_view_user_data(user):
        flash('You do not have access to view user data logs.', 'danger')
        return redirect(url_for('main.settings'))

    all_rows, filters = _load_users_summary(request.args)

    total_users = len(all_rows)
    total_activities = sum(int(row.get('activity_count') or 0) for row in all_rows)
    usage_total_seconds = sum(int(row.get('usage_seconds') or 0) for row in all_rows)
    avg_usage_seconds = int(usage_total_seconds / total_users) if total_users else 0
    today_date = datetime.now().date()
    active_today_count = sum(1 for row in all_rows if row.get('last_activity_at') and row['last_activity_at'].date() == today_date)
    active_now_count = sum(1 for row in all_rows if row.get('status_key') == 'active')

    page = int(filters['page'])
    per_page = int(filters['per_page'])
    total_pages = max(1, (total_users + per_page - 1) // per_page)
    if page > total_pages:
        page = total_pages
    start_index = (page - 1) * per_page
    end_index = start_index + per_page
    users_summary = all_rows[start_index:end_index]

    query_base = {
        'preset': filters['preset'],
        'date_from': filters['date_from'],
        'date_to': filters['date_to'],
        'q': filters['q'],
        'sort_by': filters['sort_by'],
        'sort_order': filters['sort_order'],
        'per_page': per_page,
    }
    clean_query_base = {k: v for k, v in query_base.items() if str(v) != ''}

    def _page_url(target_page: int) -> str:
        return url_for('main.users_data', **clean_query_base, page=target_page)

    def _sort_url(column: str) -> str:
        next_order = 'asc'
        if filters['sort_by'] == column and filters['sort_order'] == 'asc':
            next_order = 'desc'
        sort_query = {k: v for k, v in clean_query_base.items() if k not in {'sort_by', 'sort_order'}}
        return url_for('main.users_data', **sort_query, page=1, sort_by=column, sort_order=next_order)

    restore_batch = str(request.args.get('restore_batch') or '').strip()
    restore_batch_count = _get_soft_delete_batch_count(restore_batch)

    export_query = urlencode({k: v for k, v in clean_query_base.items() if k != 'page'})

    return render_template(
        'users_data.html',
        user=user,
        users_summary=users_summary,
        page=page,
        total_pages=total_pages,
        total_users=total_users,
        total_activities=total_activities,
        active_today_count=active_today_count,
        active_now_count=active_now_count,
        avg_usage_duration=_format_duration_full_seconds(avg_usage_seconds),
        usage_total_duration=_format_duration_full_seconds(usage_total_seconds),
        sort_by=filters['sort_by'],
        sort_order=filters['sort_order'],
        sort_urls={
            'user_email': _sort_url('user_email'),
            'brand_name': _sort_url('brand_name'),
            'activity_count': _sort_url('activity_count'),
            'first_activity_at': _sort_url('first_activity_at'),
            'last_activity_at': _sort_url('last_activity_at'),
            'usage_seconds': _sort_url('usage_seconds'),
            'top_option_count': _sort_url('top_option_count'),
        },
        filters=filters,
        per_page_options=[10, 25, 50, 100],
        prev_url=_page_url(page - 1) if page > 1 else '',
        next_url=_page_url(page + 1) if page < total_pages else '',
        export_url=url_for('main.users_data_export') + (f'?{export_query}' if export_query else ''),
        restore_batch=restore_batch,
        restore_batch_count=restore_batch_count,
        query_base=clean_query_base,
    )


@main_bp.route('/settings/users-data/export')
@login_required
def users_data_export():
    user = get_current_user()
    if not can_view_user_data(user):
        abort(403)

    rows, _ = _load_users_summary(request.args)

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            'ID',
            'Email',
            'Phone Number',
            'Brand Name',
            'Status',
            'Total Activities',
            'Top Action',
            'Top Action Count',
            'First Activity',
            'Last Activity',
            'Usage Duration',
            'Usage Duration (Exact)',
        ]
    )
    for row in rows:
        writer.writerow(
            [
                row.get('user_code'),
                row.get('user_email'),
                row.get('phone_number'),
                row.get('brand_name'),
                str(row.get('status_key') or '').title(),
                row.get('activity_count'),
                row.get('top_option_used'),
                row.get('top_option_count'),
                row.get('first_activity_at'),
                row.get('last_activity_at'),
                row.get('usage_duration'),
                row.get('usage_duration_full'),
            ]
        )

    csv_data = output.getvalue()
    output.close()
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return Response(
        csv_data,
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename=users_data_{ts}.csv'},
    )


@main_bp.route('/settings/users-data/restore', methods=['POST'])
@login_required
def users_data_restore():
    user = get_current_user()
    if not can_view_user_data(user):
        abort(403)

    batch_id = (request.form.get('batch_id') or '').strip()
    if not batch_id:
        flash('Restore batch is missing.', 'warning')
        return redirect(url_for('main.users_data'))

    restored_count = 0
    with transaction() as connection:
        with connection.cursor() as cursor:
            _ensure_soft_delete_table(cursor)
            cursor.execute(
                """
                SELECT log_data
                FROM user_data_logs_deleted
                WHERE delete_batch_id = %s
                ORDER BY id ASC
                """,
                (batch_id,),
            )
            archived_rows = cursor.fetchall() or []
            if not archived_rows:
                flash('No archived records found for restore.', 'warning')
                return redirect(url_for('main.users_data'))

            for item in archived_rows:
                log_data = item.get('log_data') or {}
                cursor.execute(
                    """
                    INSERT INTO user_data_logs (
                        user_identity_id,
                        member_visit_id,
                        user_email,
                        phone_number,
                        brand_name,
                        option_used,
                        file_name,
                        file_path,
                        created_at,
                        updated_at
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                    """,
                    (
                        log_data.get('user_identity_id'),
                        log_data.get('member_visit_id'),
                        log_data.get('user_email'),
                        log_data.get('phone_number'),
                        log_data.get('brand_name'),
                        log_data.get('option_used'),
                        log_data.get('file_name'),
                        log_data.get('file_path'),
                        log_data.get('created_at') or datetime.now(),
                    ),
                )
                restored_count += 1

            cursor.execute(
                "DELETE FROM user_data_logs_deleted WHERE delete_batch_id = %s",
                (batch_id,),
            )

    flash(f'Restored {restored_count} activity log(s).', 'success')
    return redirect(url_for('main.users_data'))


@main_bp.route('/settings/users-data/activity')
@login_required
def users_data_activity():
    user = get_current_user()
    if not can_view_user_data(user):
        flash('You do not have access to view user data logs.', 'danger')
        return redirect(url_for('main.settings'))

    email = (request.args.get('email') or '').strip().lower()
    visit_id = (request.args.get('visit_id') or '').strip()
    restore_batch = (request.args.get('restore_batch') or '').strip()
    if not email:
        flash('User activity details were not provided.', 'warning')
        return redirect(url_for('main.users_data'))
    if _is_management_account_email(email):
        flash('Admin account activity is hidden from this view.', 'warning')
        return redirect(url_for('main.users_data'))

    if visit_id.startswith('legacy-') and visit_id[7:].isdigit():
        logs = [
            SimpleNamespace(**row)
            for row in fetch_all(
                """
                SELECT *
                FROM user_data_logs
                WHERE id = %s AND lower(user_email) = %s
                  AND COALESCE(option_used, '') NOT IN ('Login', 'Logout')
                ORDER BY created_at DESC, id DESC
                """,
                (int(visit_id[7:]), email),
            )
        ]
    elif visit_id:
        logs = [
            SimpleNamespace(**row)
            for row in fetch_all(
                """
                SELECT *
                FROM user_data_logs
                WHERE lower(user_email) = %s AND COALESCE(member_visit_id, '') = %s
                  AND COALESCE(option_used, '') NOT IN ('Login', 'Logout')
                ORDER BY created_at DESC, id DESC
                """,
                (email, visit_id),
            )
        ]
    else:
        logs = [
            SimpleNamespace(**row)
            for row in fetch_all(
                """
                SELECT *
                FROM user_data_logs
                WHERE lower(user_email) = %s
                  AND COALESCE(option_used, '') NOT IN ('Login', 'Logout')
                ORDER BY created_at DESC, id DESC
                """,
                (email,),
            )
        ]

    if not logs:
        flash('No activity found for selected user.', 'warning')
        return redirect(url_for('main.users_data'))

    identity = fetch_one(
        """
        SELECT id, brand_name, latest_phone
        FROM user_identity
        WHERE lower(user_email) = %s
        LIMIT 1
        """,
        (email,),
    ) or {}

    activity_timestamps = [log.created_at for log in logs if log.created_at is not None]
    first_activity_at = min(activity_timestamps) if activity_timestamps else None
    last_activity_at = max(activity_timestamps) if activity_timestamps else None

    option_counts: dict[str, int] = {}
    for log in logs:
        key = str(log.option_used or 'Unknown')
        option_counts[key] = option_counts.get(key, 0) + 1
    sorted_option_counts = sorted(option_counts.items(), key=lambda item: item[1], reverse=True)

    mom_dashboard_log = next(
        (log for log in logs if _normalize_option_key(log.option_used) == 'mom dashboard'),
        None,
    )
    mom_business_log = next(
        (log for log in logs if _normalize_option_key(log.option_used) == 'mom business report'),
        None,
    )
    ads_report_available = mom_dashboard_log is not None and mom_business_log is not None

    filtered_logs = [
        log
        for log in logs
        if _normalize_option_key(log.option_used) not in {'mom dashboard', 'mom business report'}
    ]

    ads_report_created_at = None
    if mom_dashboard_log is not None and mom_business_log is not None:
        dashboard_created = mom_dashboard_log.created_at or datetime.min
        business_created = mom_business_log.created_at or datetime.min
        ads_report_created_at = max(dashboard_created, business_created)

    return render_template(
        'users_data_activity.html',
        user=user,
        user_code=int(identity.get('id') or 0),
        brand_name=str(identity.get('brand_name') or ''),
        member_visit_id=visit_id,
        user_email=email,
        phone_number=str(identity.get('latest_phone') or (logs[0].phone_number if logs else '')),
        total_activities=len(logs),
        first_activity_at=first_activity_at,
        last_activity_at=last_activity_at,
        usage_duration=_format_duration(first_activity_at, last_activity_at),
        option_counts=sorted_option_counts,
        logs=filtered_logs,
        ads_report_available=ads_report_available,
        mom_dashboard_log=mom_dashboard_log,
        mom_business_log=mom_business_log,
        ads_report_created_at=ads_report_created_at,
        restore_batch=restore_batch,
        restore_batch_count=_get_soft_delete_batch_count(restore_batch),
    )


@main_bp.route('/settings/users-data/activity/ads-report')
@login_required
def users_data_ads_report():
    user = get_current_user()
    if not can_view_user_data(user):
        abort(403)

    dashboard_log_id = request.args.get('dashboard_log_id', type=int)
    business_log_id = request.args.get('business_log_id', type=int)
    if not dashboard_log_id or not business_log_id:
        flash('Ads report details are missing.', 'warning')
        return redirect(url_for('main.users_data'))

    dashboard_log = _get_log_or_404(dashboard_log_id)
    business_log = _get_log_or_404(business_log_id)

    if (
        _normalize_option_key(dashboard_log.option_used) != 'mom dashboard'
        or _normalize_option_key(business_log.option_used) != 'mom business report'
    ):
        flash('Selected logs are not valid MOM report files.', 'danger')
        return redirect(url_for('main.users_data'))

    if (
        _normalize_option_key(dashboard_log.user_email) != _normalize_option_key(business_log.user_email)
        or str(dashboard_log.phone_number or '').strip() != str(business_log.phone_number or '').strip()
    ):
        flash('Selected files do not belong to the same user.', 'danger')
        return redirect(url_for('main.users_data'))

    bulk_path = _safe_log_file_path(dashboard_log)
    business_path = _safe_log_file_path(business_log)
    rows, totals = _build_ads_report_rows(bulk_path, business_path)

    return render_template(
        'users_data_ads_report.html',
        user=user,
        dashboard_log=dashboard_log,
        business_log=business_log,
        rows=rows,
        totals=totals,
    )


@main_bp.route('/settings/users-data/activity/ads-report/download')
@login_required
def users_data_ads_report_download():
    user = get_current_user()
    if not can_view_user_data(user):
        abort(403)

    dashboard_log_id = request.args.get('dashboard_log_id', type=int)
    business_log_id = request.args.get('business_log_id', type=int)
    if not dashboard_log_id or not business_log_id:
        flash('Ads report details are missing.', 'warning')
        return redirect(url_for('main.users_data'))

    dashboard_log = _get_log_or_404(dashboard_log_id)
    business_log = _get_log_or_404(business_log_id)

    if (
        _normalize_option_key(dashboard_log.option_used) != 'mom dashboard'
        or _normalize_option_key(business_log.option_used) != 'mom business report'
    ):
        flash('Selected logs are not valid MOM report files.', 'danger')
        return redirect(url_for('main.users_data'))

    if (
        _normalize_option_key(dashboard_log.user_email) != _normalize_option_key(business_log.user_email)
        or str(dashboard_log.phone_number or '').strip() != str(business_log.phone_number or '').strip()
    ):
        flash('Selected files do not belong to the same user.', 'danger')
        return redirect(url_for('main.users_data'))

    bulk_path = _safe_log_file_path(dashboard_log)
    business_path = _safe_log_file_path(business_log)
    rows, totals = _build_ads_report_rows(bulk_path, business_path)

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            'Row Labels',
            'Sum of Impressions',
            'Sum of Clicks',
            'Sum of Page Views - Total',
            'Sum of Sessions - Total',
            'Sum of Spend',
            'Sum of Sales',
            'Sum of Units Ordered',
            'Sum of Ordered Product Sales',
        ]
    )
    for row in rows:
        writer.writerow(
            [
                row['asin'],
                f"{row['impressions']:.6f}",
                f"{row['clicks']:.6f}",
                f"{row['page_views']:.6f}",
                f"{row['sessions']:.6f}",
                f"{row['spend']:.6f}",
                f"{row['sales']:.6f}",
                f"{row['units_ordered']:.6f}",
                f"{row['ordered_product_sales']:.6f}",
            ]
        )
    writer.writerow(
        [
            'Total',
            f"{totals['impressions']:.6f}",
            f"{totals['clicks']:.6f}",
            f"{totals['page_views']:.6f}",
            f"{totals['sessions']:.6f}",
            f"{totals['spend']:.6f}",
            f"{totals['sales']:.6f}",
            f"{totals['units_ordered']:.6f}",
            f"{totals['ordered_product_sales']:.6f}",
        ]
    )

    filename = f"ads_report_{dashboard_log.user_email}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
    response = Response(output.getvalue(), mimetype='text/csv')
    response.headers['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
    return response


@main_bp.route('/settings/users-data/<int:log_id>/delete-file', methods=['POST'])
@login_required
def users_data_delete_file(log_id: int):
    user = get_current_user()
    if not can_view_user_data(user):
        abort(403)

    log = _get_log_or_404(log_id)
    batch_id = uuid4().hex
    deleted_by = str(getattr(user, 'username', '') or '').strip().lower()
    with transaction() as connection:
        with connection.cursor() as cursor:
            _ensure_soft_delete_table(cursor)
            cursor.execute(
                """
                INSERT INTO user_data_logs_deleted (delete_batch_id, original_id, deleted_at, deleted_by, delete_scope, log_data)
                SELECT %s, logs.id, NOW(), %s, %s, to_jsonb(logs)
                FROM user_data_logs logs
                WHERE logs.id = %s
                """,
                (batch_id, deleted_by, 'single-file', log.id),
            )
            cursor.execute("DELETE FROM user_data_logs WHERE id = %s", (log.id,))

    flash('Activity row moved to recycle bin. You can restore it.', 'success')
    return redirect(
        url_for(
            'main.users_data_activity',
            email=log.user_email,
            visit_id=(str(getattr(log, 'member_visit_id', '') or f'legacy-{log.id}')),
            restore_batch=batch_id,
        )
    )


@main_bp.route('/settings/users-data/delete-user', methods=['POST'])
@login_required
def users_data_delete_user():
    user = get_current_user()
    if not can_view_user_data(user):
        abort(403)

    email = (request.form.get('email') or '').strip().lower()
    visit_id = (request.form.get('visit_id') or '').strip()
    if not email:
        flash('Email is required to delete user activity.', 'danger')
        return redirect(url_for('main.users_data'))
    if _is_management_account_email(email):
        flash('Admin account activity cannot be deleted from this view.', 'warning')
        return redirect(url_for('main.users_data'))

    if visit_id.startswith('legacy-') and visit_id[7:].isdigit():
        logs = [
            SimpleNamespace(**row)
            for row in fetch_all(
                """
                SELECT *
                FROM user_data_logs
                WHERE id = %s AND lower(user_email) = %s
                """,
                (int(visit_id[7:]), email),
            )
        ]
    elif visit_id:
        logs = [
            SimpleNamespace(**row)
            for row in fetch_all(
                """
                SELECT *
                FROM user_data_logs
                WHERE lower(user_email) = %s AND COALESCE(member_visit_id, '') = %s
                """,
                (email, visit_id),
            )
        ]
    else:
        logs = [
            SimpleNamespace(**row)
            for row in fetch_all(
                """
                SELECT *
                FROM user_data_logs
                WHERE lower(user_email) = %s
                """,
                (email,),
            )
        ]
    if not logs:
        flash('No user activity found for deletion.', 'warning')
        return redirect(url_for('main.users_data'))

    removed_logs = len(logs)
    referenced_files = len({str(log.file_path or '').strip() for log in logs if str(log.file_path or '').strip()})
    batch_id = uuid4().hex
    deleted_by = str(getattr(user, 'username', '') or '').strip().lower()

    with transaction() as connection:
        with connection.cursor() as cursor:
            _ensure_soft_delete_table(cursor)
            if visit_id.startswith('legacy-') and visit_id[7:].isdigit():
                cursor.execute(
                    """
                    INSERT INTO user_data_logs_deleted (delete_batch_id, original_id, deleted_at, deleted_by, delete_scope, log_data)
                    SELECT %s, logs.id, NOW(), %s, %s, to_jsonb(logs)
                    FROM user_data_logs logs
                    WHERE logs.id = %s AND lower(logs.user_email) = %s
                    """,
                    (batch_id, deleted_by, 'visit', int(visit_id[7:]), email),
                )
                cursor.execute(
                    "DELETE FROM user_data_logs WHERE id = %s AND lower(user_email) = %s",
                    (int(visit_id[7:]), email),
                )
            elif visit_id:
                cursor.execute(
                    """
                    INSERT INTO user_data_logs_deleted (delete_batch_id, original_id, deleted_at, deleted_by, delete_scope, log_data)
                    SELECT %s, logs.id, NOW(), %s, %s, to_jsonb(logs)
                    FROM user_data_logs logs
                    WHERE lower(logs.user_email) = %s AND COALESCE(logs.member_visit_id, '') = %s
                    """,
                    (batch_id, deleted_by, 'visit', email, visit_id),
                )
                cursor.execute(
                    """
                    DELETE FROM user_data_logs
                    WHERE lower(user_email) = %s AND COALESCE(member_visit_id, '') = %s
                    """,
                    (email, visit_id),
                )
            else:
                cursor.execute(
                    """
                    INSERT INTO user_data_logs_deleted (delete_batch_id, original_id, deleted_at, deleted_by, delete_scope, log_data)
                    SELECT %s, logs.id, NOW(), %s, %s, to_jsonb(logs)
                    FROM user_data_logs logs
                    WHERE lower(logs.user_email) = %s
                    """,
                    (batch_id, deleted_by, 'user', email),
                )
                cursor.execute(
                    "DELETE FROM user_data_logs WHERE lower(user_email) = %s",
                    (email,),
                )

    flash(
        f'User activity moved to recycle bin: {removed_logs} log(s), {referenced_files} file reference(s). Restore available.',
        'success',
    )
    return redirect(url_for('main.users_data', restore_batch=batch_id))


@main_bp.route('/settings/users-data/<int:log_id>/download')
@login_required
def users_data_download(log_id: int):
    user = get_current_user()
    if not can_view_user_data(user):
        abort(403)

    log = _get_log_or_404(log_id)
    file_path = _safe_log_file_path(log)
    return send_file(file_path, as_attachment=True, download_name=log.file_name)


@main_bp.route('/settings/users-data/<int:log_id>/view')
@login_required
def users_data_view(log_id: int):
    user = get_current_user()
    if not can_view_user_data(user):
        abort(403)

    log = _get_log_or_404(log_id)
    file_path = _safe_log_file_path(log)
    header, rows = _preview_file_rows(file_path)

    return render_template(
        'users_data_file_view.html',
        user=user,
        log=log,
        header=header,
        rows=rows,
    )

